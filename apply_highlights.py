# -*- coding: utf-8 -*-
"""
기준가격.xlsx에 하이라이트 배치 적용 스크립트

JSON 로그를 읽어 마감된 방 셀에 하이라이트를 안전하게 적용합니다.
- 읽기: 기준가격_하이라이트.json 로드
- 쓰기: 기준가격.xlsx에 진한 노란색(FFFF00) 하이라이트만 적용
- 완료: JSON 파일 백업 후 제거

사용법:
    py apply_highlights.py

결과:
    - 기준가격.xlsx 업데이트 (하이라이트만 추가)
    - 기준가격_하이라이트.json.bak 백업 생성
    - 기준가격_하이라이트.json 삭제
"""
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime


def apply_highlights_from_json(base_dir: str = None):
    base_dir = base_dir or os.path.dirname(__file__)
    xlsx_path = os.path.join(base_dir, "기준가격.xlsx")
    json_path = os.path.join(base_dir, "기준가격_하이라이트.json")

    # JSON 파일 확인
    if not os.path.exists(json_path):
        print(f"ℹ 적용할 하이라이트 로그가 없습니다: {json_path}")
        return 0

    # JSON 로드
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            highlights = json.load(f)
    except Exception as e:
        print(f"⚠ JSON 로드 실패: {e}")
        return 1

    if not highlights:
        print("ℹ 하이라이트 항목이 없습니다.")
        return 0

    # 엑셀 파일 열기
    try:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        applied_count = 0
        # 각 하이라이트 항목에 대해 셀 찾아서 적용
        for item in highlights:
            try:
                target_date = item.get("date")
                room_type = item.get("room_type")
                col_idx = item.get("col_idx")

                # 날짜 행 찾기
                target_row = None
                for row_idx in range(2, ws.max_row + 1):
                    try:
                        cell_val = ws.cell(row=row_idx, column=1).value
                        if cell_val:
                            import pandas as pd
                            cell_date = pd.Timestamp(cell_val).strftime("%Y-%m-%d")
                            if cell_date == target_date:
                                target_row = row_idx
                                break
                    except Exception:
                        continue

                if target_row and col_idx:
                    cell = ws.cell(row=target_row, column=col_idx)
                    cell.fill = yellow_fill
                    applied_count += 1
                    print(f"  ✓ {target_date} {room_type}(col {col_idx}) → 하이라이트 적용")
                else:
                    print(f"  ⚠ {target_date} {room_type} 셀을 찾지 못했습니다.")
            except Exception as e:
                print(f"  ⚠ 항목 적용 실패: {e}")

        # 엑셀 파일 저장
        wb.save(xlsx_path)
        print(f"\n✓ 총 {applied_count}개 하이라이트 적용 완료")

        # JSON 파일 백업 후 삭제
        backup_path = json_path + ".bak"
        try:
            import shutil
            shutil.copy(json_path, backup_path)
            os.remove(json_path)
            print(f"✓ 로그 백업: {backup_path}")
            print(f"✓ 원본 로그 삭제: {json_path}")
        except Exception as e:
            print(f"⚠ 로그 정리 실패 (수동 삭제 필요): {e}")

        return 0

    except Exception as e:
        print(f"⚠ 엑셀 파일 처리 실패: {e}")
        return 2


if __name__ == "__main__":
    exit_code = apply_highlights_from_json()
    raise SystemExit(exit_code)
