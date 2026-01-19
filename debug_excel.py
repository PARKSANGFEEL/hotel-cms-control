# -*- coding: utf-8 -*-
"""
기준가격.xlsx 파일 검증 및 디버깅 스크립트
- 현재 엑셀 파일의 실제 형식 확인
- 첫 5개 행 및 날짜 열 상세 출력
"""
import os
import pandas as pd
from openpyxl import load_workbook

base_dir = os.path.dirname(__file__)
xlsx_path = os.path.join(base_dir, "기준가격.xlsx")

print("=" * 80)
print(f"파일 경로: {xlsx_path}")
print(f"파일 존재: {os.path.exists(xlsx_path)}")
print("=" * 80)

# Pandas로 읽기
try:
    df = pd.read_excel(xlsx_path, sheet_name=0, engine='openpyxl')
    print("\n[Pandas 읽기]")
    print(f"행 수: {len(df)}")
    print(f"열 수: {len(df.columns)}")
    print(f"컬럼명: {list(df.columns)}")
    print("\n[첫 5개 행 데이터]")
    print(df.head())
    print("\n[첫 열(날짜) 상세 정보]")
    print(f"첫 열 이름: '{df.columns[0]}'")
    print(f"첫 열 데이터 타입: {df.iloc[:, 0].dtype}")
    for idx, val in enumerate(df.iloc[:, 0].head(15)):
        print(f"  [{idx}] {type(val).__name__}: {repr(val)}")
except Exception as e:
    print(f"⚠ Pandas 읽기 실패: {e}")

# Openpyxl로 직접 읽기
print("\n" + "=" * 80)
print("[Openpyxl 직접 읽기]")
try:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    print(f"시트명: {ws.title}")
    print(f"최대 행: {ws.max_row}")
    print(f"최대 열: {ws.max_column}")
    print("\n[첫 5개 행]")
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            row_data.append(f"'{val}'({type(val).__name__})" if val is not None else "None")
        print(f"  행 {row_idx}: {row_data[:3]}...")  # 처음 3개만 출력
except Exception as e:
    print(f"⚠ Openpyxl 읽기 실패: {e}")
