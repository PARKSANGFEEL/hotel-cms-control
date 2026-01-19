# -*- coding: utf-8 -*-
"""
호텔 CMS 제어 프로그램
방 타입별 예약 가능 수량 조절 기능
"""
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import time
import random
import config
from datetime import datetime, timedelta
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import re
import zipfile
import json
import shutil


def safe_save_excel(wb, excel_path):
    """엑셀 파일을 안전하게 저장 (백업 → 임시파일 → 교체 방식)"""
    try:
        backup_path = excel_path + '.bak'
        temp_path = excel_path + '.tmp'
        
        # 1. 기존 파일이 있으면 백업 생성
        if os.path.exists(excel_path):
            shutil.copy2(excel_path, backup_path)
        
        # 2. 임시 파일에 저장
        wb.save(temp_path)
        
        # 3. 임시 파일이 정상적으로 생성되었는지 확인
        if not os.path.exists(temp_path):
            raise Exception("임시 파일 생성 실패")
        
        # 4. 원본 파일 교체
        if os.path.exists(excel_path):
            os.remove(excel_path)
        shutil.move(temp_path, excel_path)
        
        # 5. 백업 파일 삭제 (성공 시)
        if os.path.exists(backup_path):
            os.remove(backup_path)
        
        return True
    except Exception as e:
        print(f"    ⚠ 엑셀 저장 실패: {e}")
        
        # 복구: 백업에서 원본 복원
        if os.path.exists(backup_path):
            try:
                if os.path.exists(excel_path):
                    os.remove(excel_path)
                shutil.copy2(backup_path, excel_path)
                print(f"    ✓ 백업에서 복구 완료")
            except Exception as restore_err:
                print(f"    ⚠ 백업 복구 실패: {restore_err}")
        
        # 임시 파일 정리
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
        
        return False


class HotelCMSController:
    """호텔 CMS를 제어하는 클래스"""
    
    # 한글 방 타입을 Excel 기준 타입으로 매핑
    KOREAN_TO_ROOM_TYPE = {
        '싱글': 'Single Room',
        '싱글룸': 'Single Room',
        '1인실': 'Single Room',
        '트윈': 'Twin Room',
        '트윈룸': 'Twin Room',
        '이코노미': 'Economy Double Room',  # 이코노미가 더블보다 먼저 확인됨
        '이코노미더블': 'Economy Double Room',
        '이코노미 더블': 'Economy Double Room',
        'economic': 'Economy Double Room',  # Agoda에서 사용
        '더블': 'Double Room',
        '더블룸': 'Double Room',
        '트리플': 'Triple Room',
        '트리플룸': 'Triple Room',
        '3인실': 'Triple Room',
        '3인': 'Triple Room',
        '3p': 'Triple Room',
        '패밀리': 'Family Room',
        '가족': 'Family Room',
        '5인': 'Family 5p',
        '5p': 'Family 5p',
    }

    def log_price_change(self, change_date, target_date, room_type, old_price, new_price, ota_label):
        """가격 변경 이력을 기준가격.xlsx의 '기준가격로그' 시트에 기록"""
        try:
            excel_path = os.path.join(os.path.dirname(__file__), "기준가격.xlsx")
            
            # Excel 파일 열기
            wb = load_workbook(excel_path)
            
            # '기준가격로그' 시트 확인 및 생성
            if '기준가격로그' not in wb.sheetnames:
                ws = wb.create_sheet('기준가격로그')
                # 헤더 작성
                ws.append(['변경날짜', '변경한날짜', '변경한룸', '변경전가격', '변경후가격', 'OTA'])
            else:
                ws = wb['기준가격로그']
            
            # 새 행 추가 (가격은 쉼표 형식)
            ws.append([
                change_date,  # 년월일시분초 형식
                target_date,
                room_type,
                f"{old_price:,}",  # 쉼표 포함
                f"{new_price:,}",  # 쉼표 포함
                ota_label[:40] if len(ota_label) > 40 else ota_label
            ])
            
            # 안전한 저장 (백업 → 임시파일 → 교체)
            if not safe_save_excel(wb, excel_path):
                print(f"    ⚠ 로그 저장 실패 (백업 확인 필요)")
        except Exception as e:
            print(f"    ⚠ 가격 변경 로그 기록 실패: {e}")

    def log_highlight_cells(self, target_date, closed_rooms):
        """마감된 방 정보를 JSON 로그에 기록 (나중에 배치 스크립트에서 엑셀에 적용)"""
        try:
            excel_path = os.path.join(os.path.dirname(__file__), "기준가격.xlsx")
            log_path = os.path.join(os.path.dirname(__file__), "기준가격_하이라이트.json")
            
            # 기존 JSON 파일 로드 또는 새 리스트 시작
            highlights = []
            if os.path.exists(log_path):
                try:
                    with open(log_path, 'r', encoding='utf-8') as f:
                        highlights = json.load(f)
                except Exception:
                    highlights = []
            
            # 엑셀 헤더 읽기 (읽기 전용)
            try:
                df = pd.read_excel(excel_path, sheet_name=0, engine='openpyxl')
                headers = list(df.columns)
            except Exception as e:
                print(f"  ⚠ 헤더 읽기 실패: {e}")
                headers = []
            
            # 마감 방 정보를 JSON에 추가
            for room_type in closed_rooms:
                if room_type in headers:
                    col_idx = headers.index(room_type) + 1  # 1-indexed
                    highlights.append({
                        "date": target_date,
                        "room_type": room_type,
                        "col_idx": col_idx
                    })
            
            # JSON 파일 저장
            with open(log_path, 'w', encoding='utf-8') as f:
                json.dump(highlights, f, ensure_ascii=False, indent=2)
            
            print(f"  ✓ {len(closed_rooms)}개 마감 방 정보를 하이라이트 로그에 기록")
            print(f"    → 나중에 'py apply_highlights.py'로 엑셀에 반영하세요")
        except Exception as e:
            print(f"  ⚠ 하이라이트 로그 기록 실패: {e}")

    def highlight_closed_rooms_in_excel(self, target_date, closed_rooms):
        """마감된 방 타입을 하이라이트 로그에 기록 (JSON 방식)"""
        self.log_highlight_cells(target_date, closed_rooms)

    def load_base_prices_from_excel(self, target_date=None):
        """엑셀 파일에서 특정 날짜의 기준가 로드 (방 타입별 기준가)"""
        try:
            excel_path = os.path.join(os.path.dirname(__file__), "기준가격.xlsx")
            if not os.path.exists(excel_path):
                print(f"  ⚠ 기준가격.xlsx 파일을 찾지 못했습니다: {excel_path}")
                return {}
            
            # 기준값이 될 날짜 (기본: 오늘)
            if not target_date:
                target_date = datetime.now().strftime("%Y-%m-%d")
            
            # 파일 시그니처/확장자 기반 엔진 선택 + 폴백들
            df = None
            ext = os.path.splitext(excel_path)[1].lower()
            # ZIP 시그니처 확인 (xlsx이면 파일 처음 2바이트가 PK)
            is_zip_xlsx = False
            try:
                with open(excel_path, 'rb') as f:
                    is_zip_xlsx = f.read(2) == b'PK'
            except Exception:
                pass

            # xlsx 손상 여부 사전 점검 (필수 엔트리 존재 확인)
            corrupted_xlsx = False
            if ext == '.xlsx' or is_zip_xlsx:
                try:
                    with zipfile.ZipFile(excel_path, 'r') as zf:
                        names = zf.namelist()
                        if '[Content_Types].xml' not in names:
                            corrupted_xlsx = True
                except Exception as e_zip:
                    corrupted_xlsx = True
                    print(f"  ⚠ 엑셀 파일 압축 구조 점검 실패: {e_zip}")

            # 손상 감지 시: 동일 폴더의 CSV 대체 파일 우선 시도
            if corrupted_xlsx:
                print("  ❗ 기준가격.xlsx가 손상된 것으로 보입니다. CSV 대체 또는 새 xlsx로 재저장을 권장합니다.")
                csv_alt = os.path.join(os.path.dirname(__file__), "기준가격.csv")
                if os.path.exists(csv_alt):
                    try:
                        df = pd.read_csv(csv_alt, encoding='utf-8')
                        print("  ✓ 손상 감지: CSV 대체 파일로 로드했습니다 (기준가격.csv)")
                    except Exception:
                        try:
                            df = pd.read_csv(csv_alt, encoding='cp949')
                            print("  ✓ 손상 감지: CSV(cp949) 대체로 로드했습니다 (기준가격.csv)")
                        except Exception as e_csv_alt:
                            print(f"  ⚠ CSV 대체 로드 실패: {e_csv_alt}")
                            df = None

            # 1) xlsx 추정: pandas(openpyxl) → openpyxl 직접
            if df is None and (ext == '.xlsx' or is_zip_xlsx):
                try:
                    df = pd.read_excel(excel_path, sheet_name=0, engine='openpyxl')
                except Exception as e_pd:
                    print(f"  ⚠ pandas(openpyxl)로 엑셀 로드 실패, openpyxl로 재시도: {e_pd}")
                    try:
                        wb = load_workbook(excel_path, read_only=True, data_only=True)
                        ws = wb.active
                        headers = [cell.value for cell in ws[1]]
                        rows = []
                        for r in ws.iter_rows(min_row=2, values_only=True):
                            rows.append(list(r))
                        df = pd.DataFrame(rows, columns=headers)
                    except Exception as e_ox:
                        print(f"  ⚠ openpyxl 폴백도 실패: {e_ox}")
                        df = None
            # 2) xls 추정: pandas(xlrd)
            if df is None and ext == '.xls':
                try:
                    df = pd.read_excel(excel_path, sheet_name=0, engine='xlrd')
                except Exception as e_xls:
                    print(f"  ⚠ xls 로드 실패 (xlrd 필요): {e_xls}")
                    df = None
            # 3) 마지막 폴백: CSV로 시도 (엑셀 대신 CSV가 저장되었을 가능성)
            if df is None:
                try:
                    # 동일 파일명 CSV가 있을 수 있음
                    csv_guess_paths = [
                        excel_path,  # 사용자가 확장자 없이 저장했을 가능성
                        os.path.join(os.path.dirname(__file__), "기준가격.csv"),
                    ]
                    last_err = None
                    for csv_p in csv_guess_paths:
                        if os.path.exists(csv_p):
                            try:
                                df = pd.read_csv(csv_p, encoding='utf-8')
                                print(f"  ⚠ 엑셀 대신 CSV로 로드했습니다: {os.path.basename(csv_p)}")
                                break
                            except Exception as e_utf8:
                                last_err = e_utf8
                                try:
                                    df = pd.read_csv(csv_p, encoding='cp949')
                                    print(f"  ⚠ 엑셀 대신 CSV(cp949)로 로드했습니다: {os.path.basename(csv_p)}")
                                    break
                                except Exception as e_cp949:
                                    last_err = e_cp949
                                    continue
                    if df is None and last_err:
                        raise last_err
                except Exception:
                    pass
            if df is None:
                print("  ⚠ 기준가를 로드하지 못했습니다 (파일 손상/포맷 확인 필요)\n     → Excel에서 파일을 열어 '다른 이름으로 저장'으로 새 .xlsx로 재저장하거나, '기준가격.csv'를 생성해 주세요.")
                return {}
            
            # 첫 번째 열이 날짜 열이므로 날짜 형식으로 변환
            date_col = df.iloc[:, 0]
            
            # 일치하는 날짜 찾기
            matching_row = None
            for idx, date_val in enumerate(date_col):
                try:
                    # 날짜 값을 문자열로 변환해서 비교
                    date_str = pd.Timestamp(date_val).strftime("%Y-%m-%d")
                    if date_str == target_date:
                        matching_row = df.iloc[idx]
                        break
                except Exception:
                    continue
            
            if matching_row is None:
                print(f"  ⚠ 엑셀에서 {target_date}에 해당하는 행을 찾지 못했습니다. 첫 번째 행을 사용합니다.")
                matching_row = df.iloc[0]
            
            # 방 타입과 기준가 매핑
            base_prices = {}
            for col_idx in range(1, len(df.columns)):
                room_type = df.columns[col_idx]
                try:
                    price = int(float(matching_row.iloc[col_idx]))
                    base_prices[room_type] = price
                except (ValueError, TypeError):
                    continue
            
            if base_prices:
                # 로그 포맷: 날짜, 각 방타입별 가격
                price_str = ', '.join([f"{rt}: {price:,}원" for rt, price in base_prices.items()])
                print(f"  📌 {target_date}: {price_str}")
            return base_prices
        except Exception as e:
            print(f"  ⚠ 기준가 로드 실패: {e}")
            return {}

    def auto_set_rates_by_rmo(self, start_date=None):
        """요금관리 메뉴에서 기준가를 기반으로 OTA별 요금 자동입력 (아고다=기준가, 나머지=기준가+5,000)"""
        try:
            # 날짜 설정
            if not start_date:
                start_date = (datetime.now() + timedelta(days=14)).strftime("%Y-%m-%d")
            
            # 시작일부터 +14일까지 기준가 모두 로드 (날짜별 맵)
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            date_list = [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(15)]
            base_prices_by_date = {}
            print(f"\n📅 기준가 로드 중 ({start_date} ~ {date_list[-1]})...")
            for d in date_list:
                base_prices_by_date[d] = self.load_base_prices_from_excel(d)
            if not any(base_prices_by_date.values()):
                print("  ⚠ 기준가를 로드하지 못했습니다. 이 페이지의 저장을 건너뜁니다.")
                return
            
            # 마감된 방을 추적할 리스트
            closed_rooms_list = []

            print("\n📋 요금관리 메뉴로 이동 중...")
            rate_url = "https://wingscms.com/#/app/cm/cm03_0200"
            self.driver.get(rate_url)
            time.sleep(3)
            
            # 시작일 input 찾기 및 값 입력
            try:
                date_input = self.wait.until(EC.presence_of_element_located((By.ID, "startDatePicker")))
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", date_input)
                date_input.click()
                time.sleep(0.2)
                # 안전하게 초기화 후 값 세팅
                self.driver.execute_script("arguments[0].value = '';", date_input)
                self.driver.execute_script("arguments[0].focus();", date_input)
                date_input.clear()
                date_input.send_keys(start_date)
                # input/change 이벤트 트리거
                self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", date_input)
                self.driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", date_input)
                date_input.send_keys(Keys.ENTER)
                print(f"  ✓ 시작일 입력: {start_date}")
            except Exception as e:
                print(f"  ⚠ 시작일 입력 실패: {e}")

            time.sleep(1)

            # 전체 객실 선택 (최초 1회만 수행)
            if not self.rate_rooms_selected:
                try:
                    print("  → 전체 객실 선택 중... (최초 1회)")
                    # 드롭다운 열기
                    try:
                        dropdown_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@id,'searchRoomType') and contains(@id,'button')]")))
                    except Exception:
                        dropdown_btn = None
                    
                    if dropdown_btn:
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_btn)
                        time.sleep(0.5)
                        dropdown_btn.click()
                        time.sleep(0.8)  # 드롭다운 메뉴 열릴 때까지 대기
                        print("  ✓ 드롭다운 열음")

                    # selectall 요소 찾기
                    sel_candidates = [
                        "[data-testid='selectall']",
                        "input[data-testid='selectall-checkbox']",
                        "span[data-testid='select-all-text']",
                        "#searchRoomType-option-selectall",
                    ]
                    select_all_el = None
                    for sel in sel_candidates:
                        try:
                            select_all_el = self.driver.find_element(By.CSS_SELECTOR, sel)
                            if select_all_el:
                                break
                        except Exception:
                            continue

                    if select_all_el:
                        try:
                            aria_sel = select_all_el.get_attribute("aria-selected")
                            data_sel = select_all_el.get_attribute("data-selected")
                            if aria_sel == "true" or (data_sel and data_sel != ""):
                                print("  ✓ 전체 객실 선택 (이미 선택됨)")
                            else:
                                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_all_el)
                                time.sleep(0.3)
                                try:
                                    select_all_el.click()
                                except Exception:
                                    self.driver.execute_script("arguments[0].click();", select_all_el)
                                print("  ✓ 전체 객실 선택 (클릭함)")
                                time.sleep(0.5)
                        except Exception:
                            try:
                                self.driver.execute_script("arguments[0].click();", select_all_el)
                                print("  ✓ 전체 객실 선택 (JS 강제)")
                                time.sleep(0.5)
                            except Exception as e2:
                                print(f"  ⚠ 전체 객실 선택 실패: {e2}")
                    else:
                        print("  ⚠ 전체 객실 selectall 요소를 찾지 못했습니다.")

                    # 플래그 설정: 이후 반복에서 선택 생략
                    self.rate_rooms_selected = True
                except Exception as e:
                    print(f"  ⚠ 전체 객실 선택 실패: {e}")
            else:
                print("  → 전체 객실 선택 건너뜀 (이미 선택됨)")

            time.sleep(1)

            # 조회 버튼 클릭
            try:
                # 조회 버튼 찾기 (여러 방법으로 시도)
                search_btn = None
                try:
                    search_btn = self.driver.find_element(By.ID, "searchBtn")
                except Exception:
                    try:
                        search_btn = self.driver.find_element(By.XPATH, "//button[@id='searchBtn']")
                    except Exception:
                        try:
                            search_btn = self.driver.find_element(By.XPATH, "//button[contains(@class, 'btn-primary') and .//i[contains(@class, 'search')]]")
                        except Exception:
                            pass
                
                if not search_btn:
                    print("  ⚠ 조회 버튼을 찾지 못했습니다.")
                else:
                    # 버튼이 클릭 가능할 때까지 대기
                    self.wait.until(EC.element_to_be_clickable((By.ID, "searchBtn")))
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_btn)
                    time.sleep(0.5)
                    
                    # 클릭 시도 (일반 클릭 → JS 클릭)
                    try:
                        search_btn.click()
                    except Exception:
                        self.driver.execute_script("arguments[0].click();", search_btn)
                    
                    print("  ✓ 조회 버튼 클릭")
                    print("  ⏳ 페이지 로드 대기 중...")
                    # 페이지가 제대로 로드될 때까지 충분히 기다림
                    time.sleep(5)
                    # 추가로 RMO 버튼이 나타날 때까지 대기
                self.wait.until(EC.presence_of_all_elements_located((By.XPATH, "//span[contains(.,'RMO')]")))
                print("  ✓ 페이지 로드 완료")
            except Exception as e:
                print(f"  ⚠ 조회 버튼 클릭 또는 페이지 로드 실패: {e}")

            # 각 객실별 RMO 버튼 클릭 및 요금 입력 (하위 child 행에 반영)
            processed_indices = set()  # 이미 처리한 RMO 버튼 index 추적
            total_rmo_count = 0
            iteration_count = 0
            
            while True:
                iteration_count += 1
                rmo_buttons = self.driver.find_elements(By.XPATH, "//span[contains(.,'RMO')]")
                
                if iteration_count == 1:
                    print(f"  ✓ RMO 버튼 {len(rmo_buttons)}개 발견")
                    total_rmo_count = len(rmo_buttons)
                
                if not rmo_buttons:
                    break
                
                print(f"    [Iteration {iteration_count}] RMO 버튼 {len(rmo_buttons)}개 찾음, 처리된 RMO: {len(processed_indices)}/{total_rmo_count}")
                
                found_unprocessed = False
                # **수정**: index 기반으로 미처리된 첫 번째 RMO 찾기
                for idx in range(len(rmo_buttons)):
                    if idx in processed_indices:
                        continue
                    
                    found_unprocessed = True
                    processed_indices.add(idx)
                    print(f"    [Iteration {iteration_count}] RMO #{idx+1} 처리 중")
                    
                    try:
                        rmo_btn = rmo_buttons[idx]
                        
                        # RMO 버튼이 속한 행의 parent_id 파악
                        parent_tr = rmo_btn.find_element(By.XPATH, "ancestor::tr")
                        parent_id = None
                        try:
                            parent_td = parent_tr.find_element(By.XPATH, ".//td[@id]")
                            parent_id = parent_td.get_attribute("id")
                        except Exception:
                            pass

                        if not parent_id:
                            continue

                        # 판매 상태 확인 (같은 방 타입의 판매 상태 행 찾기)
                        try:
                            # parent_tr의 바로 다음 행에서 data-field='CLOSE_YN' 찾기
                            try:
                                status_row = parent_tr.find_element(By.XPATH, "following-sibling::tr[@data-field='CLOSE_YN'][1]")
                                status_text = status_row.text.strip()
                            except:
                                # XPath 실패 시 다른 방법 시도: parent 텍스트에서 마감 여부 확인
                                parent_text = parent_tr.text.strip().lower()
                                status_text = parent_text
                            
                            # "마감" 또는 "Close" 또는 "closed" 같은 텍스트 확인
                            if '마감' in status_text or 'close' in status_text.lower() or '품절' in status_text:
                                # 마감된 방 타입 추출: RMO 방 타입 사용 (더 정확)
                                if rmo_room_type not in closed_rooms_list:  # 중복 방지
                                    closed_rooms_list.append(rmo_room_type)
                                print(f"    ⊘ {rmo_room_type}: 마감 상태 감지 - 스킵")
                                continue
                        except Exception as e:
                            # 판매 상태 행을 찾지 못하면 계속 진행
                            pass

                        print(f"    → RMO 버튼 활성화 중 (parent_id: {parent_id})...")
                        
                        # RMO 버튼을 스크롤해서 보이게 하고 클릭
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", rmo_btn)
                        time.sleep(0.1)
                        
                        # RMO 버튼 클릭 (여러 방법으로 시도)
                        try:
                            rmo_btn.click()
                        except Exception:
                            self.driver.execute_script("arguments[0].click();", rmo_btn)
                        
                        time.sleep(0.2)
                        
                        # RMO 버튼 클릭 후 입력란 활성화 확인 (최대 0.5초 대기)
                        rmo_input_row = None
                        for attempt in range(10):  # 0.05초 × 10 = 최대 0.5초
                            try:
                                rmo_input_row = parent_tr.find_element(By.XPATH, "following-sibling::tr[@data-field='RM_RA'][1]")
                                # 입력란이 실제로 활성화되었는지 확인 (disabled 속성 확인)
                                rmo_inputs_check = rmo_input_row.find_elements(By.CSS_SELECTOR, "input[type='text']:not([disabled])")
                                if rmo_inputs_check:
                                    print(f"    ✓ RMO 입력란 활성화 확인됨")
                                    break
                            except Exception:
                                pass
                            if attempt < 9:
                                time.sleep(0.05)
                            rmo_input_row = None

                        if not rmo_input_row:
                            print("    ⚠ RMO 입력 행을 찾지 못해 건너뜁니다.")
                            continue

                        rmo_inputs = rmo_input_row.find_elements(By.CSS_SELECTOR, "input[type='text']")
                        if not rmo_inputs:
                            print("    ⚠ RMO 입력란 없음, 건너뜀")
                            continue
                        
                        # RMO 행의 기준가를 엑셀에서 찾기 (날짜별로 적용)
                        parent_label = parent_tr.text.strip()

                        # 하위 child tr들 찾기 (같은 parent_id) — 전체를 가져온 후 활성화된 것만 처리
                        try:
                            child_trs = self.driver.find_elements(By.XPATH, f"//tr[contains(@class,'child-{parent_id}') and @data-field='RM_RA']")
                        except Exception:
                            child_trs = []

                        if not child_trs:
                            print(f"    ⚠ 하위 요금 행(child-{parent_id})을 찾지 못했습니다.")
                            continue
                        
                        # parent 행 텍스트로 방 타입 판별
                        def detect_room_type(raw_label: str):
                            norm = re.sub(r'\s+', ' ', raw_label.replace('-', ' ')).lower()
                            # 숫자/키워드 우선
                            if '2 bed' in norm or '2bed' in norm:
                                return 'Twin Room'
                            if '3 bed' in norm or '3bed' in norm or '3 person' in norm or '3person' in norm or '3p' in norm:
                                return 'Triple Room'
                            if '5 person' in norm or '5person' in norm or '5p' in norm:
                                return 'Family 5p'
                            for k, v in self.KOREAN_TO_ROOM_TYPE.items():
                                if k.lower() in norm or k in raw_label:
                                    return v
                            for room_type in sorted(['Single Room','Twin Room','Double Room','Triple Room','Economy Double Room','Family Room','Family 3p','Family 5p'], key=len, reverse=True):
                                if room_type.lower() in norm:
                                    return room_type
                            return None

                        rmo_room_type = detect_room_type(parent_label)

                        # Fallback: 버튼 순서로 매핑 (라벨이 'RMO\n요금세트'처럼 모호할 때)
                        if not rmo_room_type:
                            rmo_index_to_room_type = {
                                0: 'Single Room',
                                1: 'Economy Double Room',
                                2: 'Twin Room',
                                3: 'Double Room',
                                4: 'Triple Room',
                                5: 'Family Room 3 person',
                                6: 'Family Room 5 person'
                            }
                            rmo_room_type = rmo_index_to_room_type.get(idx)

                        if not rmo_room_type:
                            print(f"    ⚠ RMO #{idx+1} 방 타입 인식 실패: '{parent_label[:40]}'")
                            continue

                        print(f"    → child-{parent_id} 행 처리 중... (RMO 방 타입: {rmo_room_type})")

                        # OTA 매핑: 기준가를 기반으로 계산
                        def calc_new_val(label, base_price):
                            if base_price is None:
                                return None
                            label_lower = label.lower()
                            if 'agoda' in label_lower or '아고다' in label_lower:
                                return base_price
                            # Trip.com: 별도 산식 없이 기타 OTA와 동일 처리 (기준가 + 10,000~15,000원)
                            if 'trip.com' in label_lower:
                                random_addon = random.randint(10, 15) * 1000
                                return base_price + random_addon
                            # 그 외 모든 OTA: 기준가 + 10,000~15,000원 범위의 랜덤 값 (천원 단위)
                            random_addon = random.randint(10, 15) * 1000  # 10000, 11000, 12000, ..., 15000
                            return base_price + random_addon

                        for child_tr in child_trs:
                            try:
                                label = child_tr.text.strip()
                                inputs = child_tr.find_elements(By.CSS_SELECTOR, "input[type='text']")
                                if not inputs:
                                    continue
                                
                                # **수정**: 활성화된 입력칸이 있는지 확인
                                has_enabled_input = False
                                for inp in inputs:
                                    try:
                                        display = self.driver.execute_script("return window.getComputedStyle(arguments[0]).display;", inp)
                                        if display != 'none':
                                            has_enabled_input = True
                                            break
                                    except:
                                        pass
                                
                                # 활성화된 입력칸이 없으면 스킵
                                if not has_enabled_input:
                                    continue
                                
                                # 라벨이 없거나 너무 짧은 행은 스킵
                                if not label or len(label) < 5:
                                    continue
                                
                                # **핵심 수정**: 이 child_tr의 방 타입을 라벨에서 추출 (긴 이름부터 확인)
                                child_room_type = None
                                # 라벨 정규화: 하이픈을 공백으로, 여러 공백을 한 공백으로 (정규표현식 사용)
                                normalized_label = re.sub(r'\s+', ' ', label.replace('-', ' ')).lower()
                                normalized_label_no_comma = normalized_label.replace(',', '')
                                is_agoda_label = 'agoda' in normalized_label or '아고다' in normalized_label

                                # 풀라벨 정확 매핑 (요청된 라벨들만 명시적으로 5인으로 인식)
                                full_label_map = {
                                    'booking.com standard rate family studio': 'Family Room 5 person',
                                    'expedia room only family studio suite 2 bedrooms': 'Family Room 5 person',
                                    'agoda room only family room 5 pax': 'Family Room 5 person',
                                    'trip.com (new) room only family studio': 'Family Room 5 person',
                                    '야놀자 패밀리 스튜디오(주차불가) 패밀리 스튜디오 [주차불가]': 'Family Room 5 person',
                                }
                                if normalized_label_no_comma in full_label_map:
                                    child_room_type = full_label_map[normalized_label_no_comma]

                                # **특수 키워드 먼저 확인**: "2-bed room" → Twin Room, "3-bed room" → Triple Room 등
                                # '2 bedrooms' 같은 케이스는 제외하고 정확한 '2 bed(s)'만 일치
                                if re.search(r"\b2\s*bed(s)?\b", normalized_label) or re.search(r"\b2bed(s)?\b", normalized_label):
                                    child_room_type = 'Twin Room'
                                elif re.search(r"\b3\s*bed(s)?\b", normalized_label) or re.search(r"\b3bed(s)?\b", normalized_label):
                                    child_room_type = 'Triple Room'
                                elif '3 person' in normalized_label or '3person' in normalized_label or '3 pax' in normalized_label or '3p' in normalized_label:
                                    child_room_type = 'Family Room 3 person'
                                elif '5 person' in normalized_label or '5person' in normalized_label or '5 pax' in normalized_label or '5p' in normalized_label:
                                    child_room_type = 'Family Room 5 person'

                                # 추가 키워드: 'triple'만 있어도 트리플로 인식
                                if not child_room_type and 'triple' in normalized_label:
                                    child_room_type = 'Triple Room'

                                # 가족 pax 숫자 파싱 및 특수 케이스 처리: (family|패밀리) + (3/4/5...) pax/person/인, 'family/스튜디오'
                                if not child_room_type and ('family' in normalized_label or '패밀리' in normalized_label):
                                    # 'family studio'는 5인 객실로 간주하되, '1 bedroom'이 포함되면 5인으로 보지 않음
                                    if 'studio' in normalized_label or '스튜디오' in normalized_label:
                                        if re.search(r"\b1\s*bedroom(s)?\b", normalized_label) or 'one bedroom' in normalized_label:
                                            pass  # 1베드룸 스튜디오는 5인으로 인식하지 않음 (아래 기본 규칙 적용)
                                        else:
                                            child_room_type = 'Family Room 5 person'
                                    # 'standard family' 계열은 기본 3인으로 간주
                                    if not child_room_type and ('standard' in normalized_label or '스탠다드' in normalized_label):
                                        child_room_type = 'Family Room 3 person'
                                    m = re.search(r"(\d+)\s*(pax|person|인)", normalized_label)
                                    if m:
                                        try:
                                            n = int(m.group(1))
                                            if n >= 5:
                                                child_room_type = 'Family Room 5 person'
                                            elif n >= 4:
                                                child_room_type = 'Family Room 3 person'
                                            else:
                                                child_room_type = 'Family Room 3 person'
                                        except:
                                            pass
                                    # 숫자 없으면 기본 3인으로 처리 (교차 입력 방지)
                                    if not child_room_type:
                                        child_room_type = 'Family Room 3 person'

                                # **우선순위 1**: 한글/특수 키워드 먼저 확인 (이코노미, economic 등)
                                if not child_room_type:
                                    for korean_word, mapped_room_type in self.KOREAN_TO_ROOM_TYPE.items():
                                        if korean_word.lower() in normalized_label or korean_word in label:
                                            child_room_type = mapped_room_type
                                            break
                                
                                # **우선순위 2**: 영문 room_type 확인 (긴 이름부터)
                                if not child_room_type:
                                    room_types_to_check = sorted(
                                        ['Single Room', 'Twin Room', 'Double Room', 'Triple Room', 
                                         'Economy Double Room', 'Family Room 5 person', 'Family Room 3 person', 'Family Room', 'Family 3p', 'Family 5p'],
                                        key=len, reverse=True
                                    )
                                    
                                    for room_type in room_types_to_check:
                                        # 정규화된 라벨에서 room_type 검색
                                        if room_type.lower() in normalized_label:
                                            child_room_type = room_type
                                            break

                                # Family Room 일반 키는 3인으로 정규화
                                if child_room_type == 'Family Room':
                                    child_room_type = 'Family Room 3 person'

                                # Agoda 라벨인데 방 타입을 못 찾으면 현재 RMO 방 타입으로 매핑 (다른 OTAs는 스킵 유지)
                                # Agoda 라벨이라도 방 타입이 불명확하면 스킵을 위해 fallback 사용하지 않음
                                if is_agoda_label and not child_room_type:
                                    pass

                                # Agoda라도 방 타입이 식별됐는데 RMO와 다르면 스킵 (교차 입력 방지)
                                if child_room_type != rmo_room_type:
                                    continue

                                # **중요**: 이 child_tr의 방 타입이 RMO 방 타입과 정확히 일치해야만 입력
                                if not child_room_type:
                                    # 방 타입을 찾을 수 없으면 스킵
                                    continue

                                input_count = 0
                                changed_count = 0
                                skipped_count = 0
                                # 각 입력칸의 날짜별로 엑셀 기준가를 찾아 적용 (시작일~+14일)
                                for idx, inp in enumerate(inputs):
                                    # 날짜 식별: data-date/data-day가 있으면 사용, 없으면 인덱스로 date_list 매핑
                                    date_attr = inp.get_attribute("data-date") or inp.get_attribute("data-day")
                                    date_key = None
                                    if date_attr:
                                        date_key = str(date_attr).split()[0]
                                    elif idx < len(date_list):
                                        date_key = date_list[idx]

                                    if not date_key or date_key not in base_prices_by_date:
                                        continue

                                    # RMO 방 타입의 기준가 사용 (Family 키 정규화)
                                    lookup_room_type = rmo_room_type
                                    if rmo_room_type in ('Family 3p', 'Family Room 3p', 'Family 3', 'Family 3person'):
                                        lookup_room_type = 'Family Room 3 person'
                                    elif rmo_room_type in ('Family 5p', 'Family Room 5p', 'Family 5', 'Family 5person'):
                                        lookup_room_type = 'Family Room 5 person'

                                    if lookup_room_type not in base_prices_by_date[date_key]:
                                        continue
                                    
                                    day_base_price = base_prices_by_date[date_key][lookup_room_type]

                                    new_val = calc_new_val(label, day_base_price)
                                    if new_val is None:
                                        continue

                                    try:
                                        self.driver.execute_script("arguments[0].removeAttribute('readonly');", inp)
                                        display = self.driver.execute_script("return window.getComputedStyle(arguments[0]).display;", inp)
                                        if display == 'none':
                                            continue

                                        # 기존 가격 확인
                                        current_val_str = inp.get_attribute('value') or ''
                                        current_val_str = current_val_str.replace(',', '').strip()
                                        
                                        try:
                                            current_val = int(current_val_str) if current_val_str else 0
                                        except:
                                            current_val = 0
                                        
                                        # **변경 필요 여부 판단**
                                        needs_change = False
                                        label_lower = label.lower()
                                        is_agoda = ('agoda' in label_lower) or ('아고다' in label_lower)

                                        if is_agoda:
                                            # Agoda: 기준가와 정확히 일치해야 함 (다르면 무조건 덮어쓰기)
                                            needs_change = (current_val != day_base_price)
                                        else:
                                            # 기타 OTA(Trip 포함): 기준가+10,000~15,000 범위를 벗어나면 덮어쓰기
                                            expected_min = day_base_price + 10000
                                            expected_max = day_base_price + 15000
                                            if current_val == 0 or current_val < expected_min or current_val > expected_max:
                                                needs_change = True
                                            # 범위 안에 있으면 스킵
                                        
                                        if not needs_change:
                                            # 변경 필요 없음 - 스킵
                                            skipped_count += 1
                                            continue
                                        
                                        # 가격 변경 필요 - 로그 및 change_history 기록
                                        if current_val != 0:  # 기존 값이 있으면 로그 기록
                                            self.log_price_change(
                                                change_date=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                                target_date=date_key,
                                                room_type=rmo_room_type,
                                                old_price=current_val,
                                                new_price=new_val,
                                                ota_label=label
                                            )
                                            # change_history에도 추가 (저장 판단용)
                                            self.change_history.append({
                                                'date': date_key,
                                                'room_type': rmo_room_type,
                                                'old_price': current_val,
                                                'new_price': new_val,
                                                'ota_label': label[:40] if len(label) > 40 else label
                                            })
                                        
                                        self.driver.execute_script("arguments[0].focus();", inp)
                                        inp.clear()
                                        inp.send_keys(f"{new_val:,}")
                                        input_count += 1
                                        if current_val > 0:  # 기존 값이 있었다면 변경으로 카운트
                                            changed_count += 1
                                    except Exception as e:
                                        continue

                                if input_count > 0:
                                    # 입력 가격 유형 판단 (기준가 vs 기준가+애드온)
                                    if 'agoda' in label.lower() or '아고다' in label.lower():
                                        price_type = f"{new_val:,}원 (기준가)"
                                    else:
                                        addon = new_val - day_base_price
                                        price_type = f"{new_val:,}원 (기준가 {day_base_price:,}원 + {addon:,}원)"
                                    
                                    status_msg = f"입력 {input_count}개"
                                    if changed_count > 0:
                                        status_msg += f", 변경 {changed_count}개"
                                    if skipped_count > 0:
                                        status_msg += f", 스킵 {skipped_count}개"
                                    
                                    print(f"    ✓ {label}: {price_type} → {status_msg}" if label else f"    ✓ (공백 행): {price_type} → {status_msg}")
                                elif skipped_count > 0:
                                    print(f"    ⊙ {label}: 가격 동일 (스킵 {skipped_count}개)")
                            except Exception:
                                pass  # 개별 child 행 실패는 조용히 무시

                        time.sleep(0.1)
                        # 이 RMO 처리 완료, for 루프 종료 후 while으로 다음 RMO 처리
                        break
                    except Exception as e:
                        print(f"    ⚠ RMO 처리 중 오류: {e}")
                        continue
                
                # 각 RMO 처리 후 다음 진행
                
                # 모든 RMO가 처리되었으면 루프 종료
                if not found_unprocessed:
                    print(f"    [Iteration {iteration_count}] 미처리된 RMO 없음, 루프 종료")
                    break
                
                time.sleep(0.1)  # 다음 iteration 전 대기 (페이지 렌더링)

            # 변경사항이 있을 때만 저장
            if not self.change_history:
                print("\n⚠ 변경사항이 없습니다. 저장을 건너뜁니다.")
            else:
                print(f"\n📝 변경사항 {len(self.change_history)}건 발견 - 저장 진행")
                
                # 저장 버튼 클릭
                try:
                    print("💾 저장 버튼 찾는 중...")
                    # 여러 가지 방식으로 저장 버튼 찾기
                    save_btn = None
                    try:
                        save_btn = self.driver.find_element(By.XPATH, "//button[contains(text(),'저장') or contains(@class,'btn') and .//span[contains(text(),'저장')]]")
                    except:
                        try:
                            save_btn = self.driver.find_element(By.XPATH, "//button[@type='button' and contains(text(),'저장')]")
                        except:
                            try:
                                save_btn = self.driver.find_element(By.XPATH, "//button[contains(@class,'btn-primary') and contains(text(),'저장')]")
                            except:
                                pass
                    
                    if save_btn:
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", save_btn)
                        time.sleep(0.5)
                        try:
                            save_btn.click()
                        except:
                            self.driver.execute_script("arguments[0].click();", save_btn)
                        print("  ✓ 저장 버튼 클릭 완료")
                        time.sleep(2)
                        
                        # "저장되었습니다" 팝업의 확인 버튼 클릭
                        try:
                            confirm_button = self.wait.until(
                                EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-primary') and contains(., '확인')]"))
                            )
                            confirm_button.click()
                            print("  ✓ 저장 확인 완료")
                            time.sleep(1)
                        except Exception as e:
                            print(f"  ⚠ 확인 버튼 클릭 건너뜀: {e}")
                    else:
                        print("  ⚠ 저장 버튼을 찾지 못했습니다")
                except Exception as e:
                    print(f"  ⚠ 저장 버튼 클릭 실패: {e}")

            print("✓ 요금 자동입력 완료")
            
            # 마감된 방이 있으면 엑셀 파일 업데이트
            if closed_rooms_list:
                print(f"\n📊 마감된 방 {len(closed_rooms_list)}개에 대해 하이라이트 로그 기록 중...")
                self.highlight_closed_rooms_in_excel(start_date, closed_rooms_list)
        except Exception as e:
            print(f"❌ 요금 자동입력 전체 실패: {e}")
            import traceback
            traceback.print_exc()

    def __init__(self):
        """브라우저 초기화"""
        self.driver = None
        self.wait = None
        self.change_history = []  # (date, room_type, index, old_value, new_value)
        self.snapshot_rows = []   # 현재 요금 스냅샷 행 목록
        self.rate_rooms_selected = False  # 요금관리 페이지에서 객실 선택을 최초 1회만 수행

    def search_rooms_by_date(self):
        """조회 버튼을 눌러 해당 날짜의 내역을 조회"""
        try:
            search_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "searchBtn"))
            )
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_button)
            time.sleep(0.5)
            search_button.click()
            print("  ✓ 조회 버튼 클릭 - 방 목록 로딩 중...")
            time.sleep(3)
            return True
        except Exception as e:
            print(f"  ⚠ 조회 버튼 클릭 실패: {e}")
            return False

    def setup_driver(self):
        """크롬 드라이버 설정 및 초기화"""
        chrome_options = Options()
        
        if config.HEADLESS:
            chrome_options.add_argument('--headless')
        
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--start-maximized')
        
        # Selenium 4의 자동 드라이버 관리 사용
        self.driver = webdriver.Chrome(options=chrome_options)
        self.wait = WebDriverWait(self.driver, config.IMPLICIT_WAIT)
        
        print("✓ 브라우저 초기화 완료")
        
    def login(self, company_id=None, username=None, password=None):
        """CMS 로그인"""
        company_id = company_id or config.CMS_COMPANY_ID
        username = username or config.CMS_USERNAME
        password = password or config.CMS_PASSWORD
        
        if not username or not password:
            print("⚠ 로그인 정보가 설정되지 않았습니다. 수동으로 로그인해주세요.")
            return False
        
        try:
            print("\n🔐 로그인 확인 중...")
            
            # 먼저 이미 로그인되어 있는지 확인
            time.sleep(2)
            current_url = self.driver.current_url
            
            # URL에 #/app이 있으면 이미 로그인된 상태
            if "#/app" in current_url:
                print("✓ 이미 로그인되어 있습니다 (로그인 유지 상태)")
                return True
            
            # 로그인 폼이 보이는지 확인
            try:
                self.driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                print("  → 로그인 폼 발견, 로그인 진행...")
            except:
                # 로그인 폼이 없으면 이미 로그인된 것으로 간주
                print("✓ 로그인 폼이 없습니다 (이미 로그인된 상태)")
                return True
            
            # 로그인 필요
            print("\n로그인 중...")
            
            # 컴퍼니 ID 입력
            company_field = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[placeholder*='컴퍼니'], input[placeholder*='ID']"))
            )
            company_field.clear()
            company_field.send_keys(company_id)
            print(f"  ✓ 컴퍼니 ID 입력: {company_id}")
            
            # 사용자 ID/이메일 입력 - 두 번째 입력 필드
            username_fields = self.driver.find_elements(By.CSS_SELECTOR, "input[type='text']")
            if len(username_fields) >= 2:
                username_fields[1].clear()
                username_fields[1].send_keys(username)
            else:
                username_field = self.driver.find_element(By.CSS_SELECTOR, "input[placeholder*='사용자'], input[placeholder*='이메일']")
                username_field.clear()
                username_field.send_keys(username)
            print(f"  ✓ 사용자 ID 입력: {username}")
            
            # 비밀번호 입력
            password_field = self.driver.find_element(By.CSS_SELECTOR, "input[type='password']")
            password_field.clear()
            password_field.send_keys(password)
            print("  ✓ 비밀번호 입력 완료")
            
            # 로그인 유지 체크박스 찾아서 체크
            try:
                keep_login_checkbox = self.driver.find_element(By.ID, "loginKeepCheckbox")
                
                if not keep_login_checkbox.is_selected():
                    self.driver.execute_script("arguments[0].click();", keep_login_checkbox)
                    print("  ✓ 로그인 유지 체크")
                else:
                    print("  ✓ 로그인 유지 이미 체크됨")
            except Exception as e:
                print(f"  ⚠ 로그인 유지 체크 실패: {e}")
            
            time.sleep(0.5)
            
            # 로그인 버튼 클릭
            login_button = self.driver.find_element(By.XPATH, "//button[contains(text(), '로그인')]")
            login_button.click()
            print("  ✓ 로그인 버튼 클릭")
            
            # 로그인 완료 대기 (페이지 전환 또는 특정 요소 로드 확인)
            time.sleep(3)
            print("✓ 로그인 완료")
            return True
            
        except Exception as e:
            print(f"⚠ 자동 로그인 실패: {e}")
            print("수동으로 로그인해주세요.")
            import traceback
            traceback.print_exc()
            return False
    
    def navigate_to_cms(self):
        """CMS 페이지로 이동"""
        self.driver.get(config.CMS_URL)
        print(f"✓ CMS 페이지 접속: {config.CMS_URL}")
        time.sleep(3)  # 페이지 로드 대기

    def snapshot_current_rates_by_rmo(self, start_date: str, do_select_rooms: bool = False):
        """요금관리 화면에서 현재 RMO 요금(방 타입별) 15일치를 수집해서 스냅샷 행으로 누적"""
        try:
            # 15일 날짜 구성
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            date_list = [(start_dt + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(15)]

            # 요금관리 페이지 이동
            print("\n📋 요금관리 페이지로 이동 중 (스냅샷)...")
            rate_url = "https://wingscms.com/#/app/cm/cm03_0200"
            self.driver.get(rate_url)
            time.sleep(3)

            # 시작일 설정
            try:
                date_input = self.wait.until(EC.presence_of_element_located((By.ID, "startDatePicker")))
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", date_input)
                date_input.click()
                time.sleep(0.2)
                self.driver.execute_script("arguments[0].value = '';", date_input)
                self.driver.execute_script("arguments[0].focus();", date_input)
                date_input.clear()
                date_input.send_keys(start_date)
                self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles:true}));", date_input)
                self.driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", date_input)
                date_input.send_keys(Keys.ENTER)
                print(f"  ✓ 시작일 입력: {start_date}")
            except Exception as e:
                print(f"  ⚠ 시작일 입력 실패: {e}")

            time.sleep(1)

            # 전체 객실 선택 (옵션) 또는 최초 1회만 수행
            if do_select_rooms or not self.rate_rooms_selected:
                try:
                    print("  → 전체 객실 선택 중... (스냅샷)")
                    dropdown_btn = None
                    try:
                        dropdown_btn = self.wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@id,'searchRoomType') and contains(@id,'button')]")))
                    except Exception:
                        pass
                    if dropdown_btn:
                        self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", dropdown_btn)
                        time.sleep(0.5)
                        dropdown_btn.click()
                        time.sleep(0.8)
                    sel_candidates = [
                        "[data-testid='selectall']",
                        "input[data-testid='selectall-checkbox']",
                        "span[data-testid='select-all-text']",
                        "#searchRoomType-option-selectall",
                    ]
                    select_all_el = None
                    for sel in sel_candidates:
                        try:
                            select_all_el = self.driver.find_element(By.CSS_SELECTOR, sel)
                            if select_all_el:
                                break
                        except Exception:
                            continue
                    if select_all_el:
                        try:
                            aria_sel = select_all_el.get_attribute("aria-selected")
                            data_sel = select_all_el.get_attribute("data-selected")
                            if aria_sel == "true" or (data_sel and data_sel != ""):
                                pass
                            else:
                                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", select_all_el)
                                time.sleep(0.3)
                                try:
                                    select_all_el.click()
                                except Exception:
                                    self.driver.execute_script("arguments[0].click();", select_all_el)
                                time.sleep(0.5)
                        except Exception:
                            try:
                                self.driver.execute_script("arguments[0].click();", select_all_el)
                                time.sleep(0.5)
                            except Exception:
                                pass
                    else:
                        print("  ⚠ 전체 객실 선택 요소 없음")
                    # 플래그 설정: 이후 반복에서 선택 생략
                    self.rate_rooms_selected = True
                except Exception as e:
                    print(f"  ⚠ 전체 객실 선택 실패: {e}")

            time.sleep(1)

            # 조회 버튼 클릭
            try:
                search_btn = None
                try:
                    search_btn = self.driver.find_element(By.ID, "searchBtn")
                except Exception:
                    try:
                        search_btn = self.driver.find_element(By.XPATH, "//button[@id='searchBtn']")
                    except Exception:
                        try:
                            search_btn = self.driver.find_element(By.XPATH, "//button[contains(@class, 'btn-primary') and .//i[contains(@class, 'search')]]")
                        except Exception:
                            pass
                if search_btn:
                    self.wait.until(EC.element_to_be_clickable((By.ID, "searchBtn")))
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_btn)
                    time.sleep(0.5)
                    try:
                        search_btn.click()
                    except Exception:
                        self.driver.execute_script("arguments[0].click();", search_btn)
                    print("  ✓ 조회 버튼 클릭")
                    print("  ⏳ 페이지 로드 대기 중...")
                    time.sleep(5)
                self.wait.until(EC.presence_of_all_elements_located((By.XPATH, "//span[contains(.,'RMO')]")))
                print("  ✓ 페이지 로드 완료")
            except Exception as e:
                print(f"  ⚠ 조회 실패: {e}")

            # RMO 버튼 순회하여 방 타입별 현재 요금 수집
            rmo_buttons = self.driver.find_elements(By.XPATH, "//span[contains(.,'RMO')]")
            print(f"  ✓ RMO 버튼 {len(rmo_buttons)}개 발견 (스냅샷)")

            # 날짜별 스냅샷 테이블 초기화
            room_columns = [
                'Single Room', 'Economy Double Room', 'Twin Room', 'Double Room',
                'Triple Room', 'Family Room 3 person', 'Family Room 5 person'
            ]
            rates_by_date = {d: {col: None for col in room_columns} for d in date_list}

            for idx, rmo_btn in enumerate(rmo_buttons):
                try:
                    parent_tr = rmo_btn.find_element(By.XPATH, "ancestor::tr")
                    parent_label = parent_tr.text.strip()
                    print(f"    [RMO {idx+1}] 전체 텍스트:\n      {repr(parent_label)}")
                    
                    # parent_tr의 모든 td 텍스트를 수집 (전체 행 정보)
                    all_tds = parent_tr.find_elements(By.XPATH, ".//td")
                    td_texts = [td.text.strip() for td in all_tds if td.text.strip()]
                    print(f"      → TD 개수: {len(all_tds)}, 텍스트: {td_texts[:3]}")
                    
                    # parent_id 추출 (child-{id} 매칭용)
                    parent_id = None
                    try:
                        parent_td = parent_tr.find_element(By.XPATH, ".//td[@id]")
                        parent_id = parent_td.get_attribute("id")
                        print(f"      → parent_id: {parent_id}")
                    except Exception as pid_err:
                        print(f"      ⚠ parent_id 미추출: {pid_err}")
                        pass
                    # 방 타입 판별 (기존 로직 재사용)
                    def detect_room_type(raw_label: str):
                        norm = re.sub(r'\s+', ' ', raw_label.replace('-', ' ')).lower()
                        if '2 bed' in norm or '2bed' in norm:
                            return 'Twin Room'
                        if '3 bed' in norm or '3bed' in norm or '3 person' in norm or '3person' in norm or '3p' in norm:
                            return 'Triple Room'
                        if '5 person' in norm or '5person' in norm or '5p' in norm:
                            return 'Family 5p'
                        for k, v in self.KOREAN_TO_ROOM_TYPE.items():
                            if k.lower() in norm or k in raw_label:
                                return v
                        for room_type in sorted(['Single Room','Twin Room','Double Room','Triple Room','Economy Double Room','Family Room','Family 3p','Family 5p'], key=len, reverse=True):
                            if room_type.lower() in norm:
                                return room_type
                        return None

                    # RMO 클릭 (해당 객실의 child tr 노출)
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", rmo_btn)
                    time.sleep(0.3)
                    try:
                        rmo_btn.click()
                    except Exception:
                        self.driver.execute_script("arguments[0].click();", rmo_btn)
                    time.sleep(0.8)
                    
                    # ⭐ 이 RMO 행의 다음 tr들을 수집하되, 다음 RMO 버튼(span 'RMO' 포함)이 나올 때까지만
                    parent_tr_for_child = rmo_btn.find_element(By.XPATH, "ancestor::tr")
                    child_trs = []
                    
                    try:
                        # parent_tr의 다음 sibling tr들을 순회
                        current_sibling = parent_tr_for_child.find_element(By.XPATH, "following-sibling::tr[1]")
                        
                        while current_sibling:
                            try:
                                # data-field='RM_RA'인 경우만 수집 (다른 데이터는 무시)
                                if current_sibling.get_attribute("data-field") == "RM_RA":
                                    child_trs.append(current_sibling)
                                
                                # 다음 RMO 버튼을 만나면 중단 (span[contains(.,'RMO')])
                                try:
                                    next_rmo = current_sibling.find_element(By.XPATH, ".//span[contains(.,'RMO')]")
                                    break  # 다음 RMO를 만났으므로 중단
                                except:
                                    pass
                                
                                current_sibling = current_sibling.find_element(By.XPATH, "following-sibling::tr[1]")
                            except:
                                break
                    except Exception as e:
                        pass
                    
                    print(f"      → 해당 RMO의 child 행 {len(child_trs)}개 발견")
                    
                    if not child_trs:
                        print(f"      ⚠ child 행 없음, 스킵")
                        continue
                    
                    # 첫 번째 child tr의 텍스트에서 방 타입 찾기
                    print(f"        📋 RMO {idx+1} child 텍스트:")
                    for i, child_tr in enumerate(child_trs[:7]):  # 처음 7개만
                        try:
                            child_text = child_tr.text.strip()
                            if child_text:
                                print(f"          [{i}] {repr(child_text[:80])}")
                        except:
                            pass
                    
                    rmo_room_type = None
                    for child_idx, child_tr in enumerate(child_trs):
                        try:
                            child_text = child_tr.text.strip()
                            if not child_text or child_text == '요금':
                                continue
                            rmo_room_type = detect_room_type(child_text)
                            if rmo_room_type:
                                print(f"        ✓ child[{child_idx}]에서 방 타입 발견: {rmo_room_type}")
                                break
                        except Exception as ce:
                            pass
                    
                    if not rmo_room_type:
                        print(f"      ❌ 방 타입 판별 실패, 스킵")
                        continue
                    # Family 키 정규화
                    if rmo_room_type in ('Family 3p', 'Family Room', 'Family Room 3p', 'Family 3', 'Family 3person'):
                        lookup_room_type = 'Family Room 3 person'
                    elif rmo_room_type in ('Family 5p', 'Family Room 5p', 'Family 5', 'Family 5person', 'Family Studio', 'Twin Room') and idx == 6:  # RMO 7 특수 처리
                        lookup_room_type = 'Family Room 5 person'
                    elif rmo_room_type in ('Family 5p', 'Family Room 5p', 'Family 5', 'Family 5person'):
                        lookup_room_type = 'Family Room 5 person'
                    else:
                        lookup_room_type = rmo_room_type
                    print(f"      ✓ 방 타입: {lookup_room_type}")

                    # 입력값 파서
                    def parse_input_value(inp):
                        try:
                            val = self.driver.execute_script(
                                "return (arguments[0].value || arguments[0].getAttribute('value') || arguments[0].textContent || arguments[0].innerText) ?? '';",
                                inp,
                            )
                            if val is None:
                                return None
                            s = str(val).strip()
                            digits = re.sub(r"[^0-9]", "", s)
                            if digits == "":
                                return None
                            return int(digits)
                        except Exception:
                            return None

                    # ⭐ Agoda child tr에서 가격 추출 (Agoda = 기준가)
                    print(f"        💰 가격 수집 중... (child {len(child_trs)}개 순회)")
                    agoda_found = False
                    
                    # 먼저 Agoda child를 찾기 (room type 확인용)
                    agoda_child_idx = None
                    for child_idx, child_tr in enumerate(child_trs):
                        try:
                            label = child_tr.text.strip()
                            norm_label = re.sub(r'\s+', ' ', label.replace('-', ' ')).lower()
                            is_agoda = ('agoda' in norm_label) or ('아고다' in norm_label)
                            
                            if is_agoda:
                                agoda_child_idx = child_idx
                                print(f"        ✓ Agoda child 발견 (index={agoda_child_idx})")
                                break
                        except Exception:
                            pass
                    
                    # Agoda child를 찾았으면, parent_tr의 다음 형제(input 행)에서 값을 읽기
                    if agoda_child_idx is not None:
                        try:
                            # parent_tr_for_child의 다음 형제 행들을 순회해서 input 행 찾기
                            input_row = None
                            current_row = parent_tr_for_child
                            for _ in range(10):  # 최대 10개 형제 행 검색
                                try:
                                    next_row = current_row.find_element(By.XPATH, "following-sibling::tr[1]")
                                    # data-field 확인
                                    data_field = next_row.get_attribute("data-field")
                                    if data_field == "RM_RA":  # 입력 행일 가능성
                                        inputs = next_row.find_elements(By.CSS_SELECTOR, "input[type='text']")
                                        if inputs:
                                            input_row = next_row
                                            print(f"        ✓ 입력 행 발견! input {len(inputs)}개")
                                            break
                                    current_row = next_row
                                except:
                                    break
                            
                            if input_row:
                                inputs = input_row.find_elements(By.CSS_SELECTOR, "input[type='text']")
                                print(f"        → 입력창 {len(inputs)}개에서 값 읽기")
                                for i, inp in enumerate(inputs):
                                    try:
                                        date_key = date_list[i] if i < len(date_list) else None
                                        if not date_key:
                                            continue
                                        val = parse_input_value(inp)
                                        if val is not None:
                                            rates_by_date[date_key][lookup_room_type] = val
                                            print(f"            [{date_key}] {lookup_room_type}: {val} ✓")
                                        else:
                                            print(f"            [{date_key}] 값 없음")
                                    except Exception:
                                        pass
                                agoda_found = True
                        except Exception as e:
                            print(f"        ⚠ 입력 행 검색 실패: {e}")
                    
                    if not agoda_found:
                        print(f"        ⚠ Agoda 또는 입력 행을 찾지 못함")

                    # Agoda 행이 없으면, 첫 번째 OTA 행의 값을 사용 (보조)
                    if not agoda_found:
                        for child_tr in child_trs:
                            try:
                                inputs = child_tr.find_elements(By.CSS_SELECTOR, "input[type='text']")
                                if not inputs:
                                    continue
                                for i, inp in enumerate(inputs):
                                    try:
                                        date_attr = inp.get_attribute("data-date") or inp.get_attribute("data-day")
                                        date_key = None
                                        if date_attr:
                                            date_key = str(date_attr).split()[0]
                                        elif i < len(date_list):
                                            date_key = date_list[i]
                                        if not date_key:
                                            continue
                                        val = parse_input_value(inp)
                                        if date_key in rates_by_date and val is not None and rates_by_date[date_key][lookup_room_type] is None:
                                            rates_by_date[date_key][lookup_room_type] = val
                                    except Exception:
                                        continue
                                break
                            except Exception:
                                continue
                except Exception:
                    continue

            # 수집 결과를 행 형태로 누적 (기준가격.xlsx 형식)
            rows_added = 0
            for d in date_list:
                row = {'날짜': d}
                for col in room_columns:
                    row[col] = rates_by_date[d].get(col)
                self.snapshot_rows.append(row)
                rows_added += 1
            print(f"  ✓ 스냅샷 누적: {rows_added}행 추가 (총 {len(self.snapshot_rows)}행)")

        except Exception as e:
            print(f"❌ 스냅샷 수집 실패: {e}")
            try:
                import traceback
                print(f"📌 상세 오류:\n{traceback.format_exc()}")
                traceback.print_exc()
            except Exception:
                pass

    def save_snapshot_to_excel(self, output_path: str = None):
        """누적된 스냅샷 행을 기준가격과 같은 헤더로 엑셀에 저장 (쉼표 포맷, 금토 빨강)"""
        try:
            if not self.snapshot_rows:
                print("  ⚠ 저장할 스냅샷 데이터가 없습니다.")
                return False
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment
            from datetime import datetime
            
            df = pd.DataFrame(self.snapshot_rows)
            # 컬럼 순서 정렬
            columns = ['날짜', 'Single Room', 'Economy Double Room', 'Twin Room', 'Double Room', 'Triple Room', 'Family Room 3 person', 'Family Room 5 person']
            df = df.reindex(columns=columns)
            # 기본 경로
            output_path = output_path or os.path.join(os.path.dirname(__file__), '현재가격.xlsx')
            print(f"  📝 DataFrame 생성: {len(df)}행 × {len(columns)}컬럼")
            print(f"  🔍 데이터 샘플:\n{df.head(3)}")
            # 기존 파일 삭제 (권한 오류 방지)
            try:
                if os.path.exists(output_path):
                    print(f"  🗑️ 기존 파일 삭제 시작: {output_path}")
                    os.remove(output_path)
                    print(f"  ✓ 기존 파일 삭제 완료")
            except Exception as e:
                print(f"  ⚠ 기존 파일 삭제 실패: {e}")
            
            # 임시 저장
            print(f"  💾 엑셀 저장 시작...")
            df.to_excel(output_path, index=False)
            
            # openpyxl로 스타일링 적용
            wb = load_workbook(output_path)
            ws = wb.active
            
            # 빨간색 폰트 (금토일용)
            red_font = Font(color="FF0000")  # 빨간색
            
            # 각 행에 대해 스타일 적용
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                # 날짜 열 (A열)에서 요일 확인
                date_cell = ws[f'A{row_idx}']
                try:
                    date_str = str(date_cell.value)
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                    weekday = date_obj.weekday()  # 0=월, 4=금, 5=토
                    is_weekend = weekday in (4, 5)  # 금요일(4) 또는 토요일(5)
                except:
                    is_weekend = False
                
                # 각 셀에 대해 포맷 적용
                for col_idx, cell in enumerate(row, start=1):
                    if col_idx > 1:  # 숫자 컬럼 (B열 이후)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            # 숫자에 쉼표 추가
                            cell.number_format = '#,##0'
                            # 금토일 빨간색
                            if is_weekend:
                                cell.font = red_font
                    elif is_weekend and col_idx == 1:  # 날짜 열도 빨강 처리
                        cell.font = red_font
            
            # 열 너비 조정
            ws.column_dimensions['A'].width = 15
            for col_idx in range(2, len(columns) + 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 18
            
            # 안전한 저장 (백업 → 임시파일 → 교체)
            if not safe_save_excel(wb, output_path):
                print(f"  ⚠ 스냅샷 저장 실패 (백업 확인 필요)")
                return False
            
            print(f"  ✓ 스냅샷 저장 완료: {output_path} (행 {len(df)})")
            print(f"  ✓ 숫자 포맷: 쉼표 추가 ✓ 금요일/토요일: 빨간색")
            return True
        except Exception as e:
            print(f"❌ 스냅샷 저장 실패: {e}")
            try:
                import traceback
                traceback.print_exc()
            except Exception:
                pass
            return False
    
    def navigate_to_inventory_page(self, date_str=None, do_select_rooms=True):
        """인벤토리 관리_객실별 페이지로 이동"""
        try:
            print("\n📋 인벤토리 관리 페이지로 이동 중...")
            # 직접 URL로 이동
            inventory_url = "https://wingscms.com/#/app/cm/cm03_0300"
            self.driver.get(inventory_url)
            print(f"  ✓ 인벤토리 관리_객실별 페이지 이동: {inventory_url}")
            time.sleep(3)  # 페이지 로드 대기

            # 입력받은 시작일이 없으면 오늘 날짜로 셋팅
            if not date_str or str(date_str).strip() == "":
                date_str = datetime.now().strftime("%Y-%m-%d")
            self.set_date(date_str)
            # 호텔 객실 선택 및 필터 설정은 최초 1회만
            if do_select_rooms:
                self.select_all_rooms()
            return True
        except Exception as e:
            print(f"❌ 페이지 이동 실패: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def set_date(self, date_str):
        """날짜 설정 (형식: YYYY-MM-DD)"""
        try:
            print(f"\n📅 날짜 설정 중: {date_str}")
            
            # sweet-overlay나 modal이 있으면 닫기
            try:
                # overlay가 있는지 확인
                overlay = self.driver.find_element(By.CSS_SELECTOR, ".sweet-overlay")
                if overlay.is_displayed():
                    print("  → sweet-overlay 발견, 닫는 중...")
                    # ESC 키로 닫기 시도
                    from selenium.webdriver.common.keys import Keys
                    self.driver.find_element(By.TAG_NAME, 'body').send_keys(Keys.ESCAPE)
                    time.sleep(0.5)
                    # 또는 JavaScript로 강제로 제거
                    self.driver.execute_script("""
                        var overlay = document.querySelector('.sweet-overlay');
                        if (overlay) overlay.style.display = 'none';
                        var alert = document.querySelector('.sweet-alert');
                        if (alert) alert.style.display = 'none';
                    """)
                    time.sleep(0.5)
                    print("  ✓ overlay 제거 완료")
            except:
                pass  # overlay가 없으면 무시
            
            # 날짜 입력 필드 찾기
            date_input = self.wait.until(
                EC.presence_of_element_located((By.ID, "startDatePicker"))
            )
            
            # JavaScript로 직접 클릭 (overlay 우회)
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", date_input)
            time.sleep(0.3)
            try:
                date_input.click()
            except:
                # 일반 클릭 실패 시 JavaScript 클릭
                self.driver.execute_script("arguments[0].click();", date_input)
            
            print("  ✓ 달력 열기")
            time.sleep(1)
            

            # 입력받은 날짜 파싱
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            target_year = str(dt.year)
            target_month = dt.strftime("%B")  # 영어 월명
            target_day = str(dt.day)

            # 달력 네비게이션으로 목표 년월로 이동
            max_clicks = 50
            clicks = 0
            print(f"  달력을 {target_year}년 {target_month}로 이동 중...")
            while clicks < max_clicks:
                try:
                    current_month_year = self.driver.find_element(
                        By.CSS_SELECTOR,
                        ".react-datepicker__current-month"
                    ).text
                    print(f"    현재: {current_month_year}")
                    if target_month in current_month_year and target_year in current_month_year:
                        print(f"  ✓ 목표 도달: {current_month_year}")
                        break
                    next_button = self.driver.find_element(
                        By.CSS_SELECTOR,
                        ".react-datepicker__navigation--next"
                    )
                    next_button.click()
                    time.sleep(0.5)
                    clicks += 1
                except Exception as e:
                    print(f"  ⚠ 네비게이션 중 오류: {e}")
                    break

            # 해당 일(day) 클릭
            time.sleep(0.5)
            day_element = self.wait.until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    f"//div[contains(@class, 'react-datepicker__day') and not(contains(@class, 'outside-month')) and text()='{target_day}']"
                ))
            )
            day_element.click()
            print(f"  ✓ {date_str} 날짜 선택 완료")
            time.sleep(1)
            
            return True
            
        except Exception as e:
            print(f"⚠ 날짜 설정 실패: {e}")
            print("수동으로 날짜를 선택해주세요.")
            import traceback
            traceback.print_exc()
            return False
    
    def select_all_rooms(self):
        """Single Room, Twin Room, Triple Room만 선택"""
        try:
            print("\n🏨 호텔 객실 선택 중...")
            
            # 드롭다운 버튼 클릭
            dropdown_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "hotelRoomSearch__button__button"))
            )
            dropdown_button.click()
            print("  ✓ 객실 선택 드롭다운 열기")
            time.sleep(1.5)
            
            # 먼저 모든 옵션 찾기 (전체 체크 해제용)
            all_options = self.driver.find_elements(By.XPATH, "//div[@role='option' and contains(@id, 'hotelRoomSearch-option-')]")
            print(f"  → 전체 옵션 {len(all_options)}개 찾음")
            
            # 모든 옵션 체크 해제
            for option in all_options:
                is_selected = option.get_attribute("aria-selected")
                data_selected = option.get_attribute("data-selected")
                
                # 체크되어 있으면 클릭하여 해제
                if is_selected == "true" or (data_selected and data_selected != ""):
                    option_text = option.text
                    self.driver.execute_script("arguments[0].click();", option)
                    print(f"  → '{option_text}' 체크 해제")
                    time.sleep(0.3)
            
            time.sleep(1)
            
            # Single Room, Twin Room, Double Room, Triple Room 선택
            target_rooms = ["Single Room", "Twin Room", "Double Room", "Triple Room"]
            
            for room_name in target_rooms:
                try:
                    # 옵션 찾기 (텍스트로)
                    room_option = self.driver.find_element(
                        By.XPATH,
                        f"//div[@role='option' and contains(@id, 'hotelRoomSearch-option-') and text()='{room_name}']"
                    )
                    
                    # 체크 상태 확인
                    is_selected = room_option.get_attribute("aria-selected")
                    data_selected = room_option.get_attribute("data-selected")
                    
                    # 체크되어 있지 않으면 클릭
                    if is_selected != "true" and (not data_selected or data_selected == ""):
                        self.driver.execute_script("arguments[0].click();", room_option)
                        print(f"  ✓ {room_name} 선택")
                        time.sleep(0.5)
                    else:
                        print(f"  ✓ {room_name} 이미 선택됨")
                        
                except Exception as e:
                    print(f"  ⚠ {room_name} 선택 실패: {e}")
            
            time.sleep(1)
            
            # 조회 버튼 클릭
            search_button = self.wait.until(
                EC.element_to_be_clickable((By.ID, "searchBtn"))
            )
            search_button.click()
            print("  ✓ 조회 버튼 클릭 - 방 목록 로딩 중...")
            time.sleep(3)  # 방 목록이 로드될 때까지 대기
            
            print("✅ Single Room, Twin Room, Triple Room 목록이 표시되었습니다!")
            
            # 필터 설정 (필수)
            self.apply_filter()
            
            return True
            
        except Exception as e:
            print(f"⚠ 객실 선택 자동화 실패: {e}")
            print("수동으로 객실을 선택하고 조회 버튼을 눌러주세요.")
            import traceback
            traceback.print_exc()
            return False
    
    def apply_filter(self):
        """필터에서 판매가능객실만 선택 (재시도 로직 포함)"""
        print("\n🔍 필터 설정 시도 중...")
        
        max_retries = 3
        for attempt in range(max_retries):
            try:
                if attempt > 0:
                    print(f"  ⟳ 재시도 {attempt}/{max_retries-1}...")
                    time.sleep(3)
                
                # 1단계: 필터 아이콘 클릭하여 사이드 패널 열기
                try:
                    filter_button = self.wait.until(
                        EC.element_to_be_clickable((By.CLASS_NAME, "filter-ico"))
                    )
                except:
                    filter_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//span[@class='filter-ico']"))
                    )
                
                self.driver.execute_script("arguments[0].scrollIntoView(true);", filter_button)
                time.sleep(1)
                self.driver.execute_script("arguments[0].click();", filter_button)
                print("  ✓ 필터 패널 열기")
                time.sleep(3)  # 패널이 완전히 열릴 때까지 대기
                
                # 2단계: 노출정보 드롭다운 클릭
                exposure_dropdown = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "COMN_CN__button__button"))
                )
                self.driver.execute_script("arguments[0].click();", exposure_dropdown)
                print("  ✓ 노출정보 드롭다운 열기")
                time.sleep(2)
                
                # 3단계: 모든 체크박스 해제 후 "판매가능객실"만 체크
                # 먼저 모든 옵션 찾기
                all_options = self.driver.find_elements(By.XPATH, "//div[@role='option' and contains(@id, 'COMN_CN-option-')]")
                print(f"  → 전체 옵션 {len(all_options)}개 찾음")
                
                # 모든 옵션 체크 해제
                for option in all_options:
                    is_selected = option.get_attribute("aria-selected")
                    data_selected = option.get_attribute("data-selected")
                    
                    # 체크되어 있으면 클릭하여 해제
                    if is_selected == "true" or (data_selected and data_selected != ""):
                        option_text = option.text
                        self.driver.execute_script("arguments[0].click();", option)
                        print(f"  → '{option_text}' 체크 해제")
                        time.sleep(0.3)
                
                time.sleep(1)
                
                # "판매가능객실"만 체크
                sales_room_option = self.wait.until(
                    EC.presence_of_element_located((By.ID, "COMN_CN-option-0"))
                )
                
                is_selected = sales_room_option.get_attribute("aria-selected")
                data_selected = sales_room_option.get_attribute("data-selected")
                
                if is_selected != "true" and (not data_selected or data_selected == ""):
                    self.driver.execute_script("arguments[0].click();", sales_room_option)
                    print("  ✓ 판매가능객실 체크")
                    time.sleep(2)
                else:
                    print("  ✓ 판매가능객실 이미 체크됨")
                    time.sleep(1)
                
                # 4단계: 검색 버튼 클릭 (여러 방법 시도)
                search_button = None
                try:
                    # 방법 1: 텍스트로 찾기
                    search_button = self.wait.until(
                        EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-primary') and contains(text(), '검색')]"))
                    )
                    print("  → 검색 버튼 찾음 (텍스트)")
                except:
                    try:
                        # 방법 2: 아이콘 클래스로 찾기
                        search_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-primary')]//i[contains(@class, 'pe-7s-search')]"))
                        )
                        # 부모 버튼 요소로 이동
                        search_button = search_button.find_element(By.XPATH, "..")
                        print("  → 검색 버튼 찾음 (아이콘)")
                    except:
                        # 방법 3: w90 클래스로 찾기
                        search_button = self.wait.until(
                            EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'w90') and contains(@class, 'btn-primary')]"))
                        )
                        print("  → 검색 버튼 찾음 (클래스)")
                
                # 스크롤하여 버튼이 보이도록
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", search_button)
                time.sleep(1)
                
                # 여러 방법으로 클릭 시도
                try:
                    search_button.click()
                    print("  ✓ 검색 버튼 클릭 (일반 클릭)")
                except:
                    self.driver.execute_script("arguments[0].click();", search_button)
                    print("  ✓ 검색 버튼 클릭 (JavaScript)")
                
                time.sleep(5)  # 결과 로딩 충분히 대기
                
                print("✅ 필터 적용 완료 - 판매가능객실만 표시됨!")
                return True
                
            except Exception as e:
                print(f"  ✗ 시도 실패: {e}")
                if attempt == max_retries - 1:
                    print(f"\n❌ 필터 적용 {max_retries}회 시도 후 실패")
                    raise Exception(f"필터 적용 실패 - 입력칸이 생성되지 않아 작업을 계속할 수 없습니다: {e}")
                continue
        
        return False
    
    def calculate_available_rooms(self, room_type, remaining, booked, max_count):
        """
        판매가능객실 수량 계산
        
        Args:
            room_type: 'SINGLE', 'TWIN', 'TRIPLE'
            remaining: 잔여 수
            booked: 예약 수
            max_count: 최대 수량
        
        Returns:
            판매가능객실 수량 (None이면 빈칸 = 모두 오픈), 또는 'ALERT:메시지'
        """
        print(f"[정책진단] room_type={room_type}, remaining={remaining}, booked={booked}, max_count={max_count}")

        if room_type in ('SINGLE', 'TWIN', 'DOUBLE'):
            # 더블룸: 예약 3개 이상이면 모두 오픈(빈칸), 아니면 예약+2~3개(랜덤)
            if room_type == 'DOUBLE':
                if booked >= 3:
                    print(f"[정책결과] => None (더블룸 예약 3 이상, 모두 오픈)")
                    return None
                else:
                    val = booked + 2
                    print(f"[정책결과] => {val} (더블룸 예약+2 고정)")
                    return val

        if room_type == 'SINGLE':
            if remaining <= 4:
                print(f"[정책결과] => None (싱글룸 잔여 4 이하, 모두 오픈)")
                return None
            val = booked + 4
            print(f"[정책결과] => {val} (싱글룸 예약+4 고정)")
            return val

        elif room_type == 'TWIN':
            if remaining <= 4:
                print(f"[정책결과] => None (트윈룸 잔여 4 이하, 모두 오픈)")
                return None
            if remaining >= 3 and booked >= 6:
                print(f"[정책결과] => {booked + 2} (트윈룸 잔여 3이상 & 예약 6이상, 예약+2)")
                return booked + 2
            if remaining >= 3 and booked < 6:
                print(f"[정책결과] => {booked + 4} (트윈룸 잔여 3이상 & 예약 6미만, 예약+4)")
                return booked + 4

        elif room_type == 'TRIPLE':
            if booked >= 5:
                print(f"[정책결과] => ALERT:트리플룸 예약 {booked}건 - 수동 확인 필요")
                return f"ALERT:트리플룸 예약 {booked}건 - 수동 확인 필요"
            elif booked == 4:
                print(f"[정책결과] => {booked + 1} (예약=4, +1)")
                return booked + 1  # 5
            else:
                print(f"[정책결과] => {booked + 2} (예약<4, +2)")
                return booked + 2

        val = random.randint(2, 3)
        print(f"[정책결과] => {val} (기본값)")
        return val
    
    def set_room_availability_by_date(self):
        from datetime import datetime, timedelta
        run_date = datetime.now().strftime('%Y-%m-%d')
        # 현재 날짜 추출 (달력에서 선택된 값)
        try:
            date_input = self.driver.find_element(By.ID, "startDatePicker")
            current_date = date_input.get_attribute("value")
        except:
            current_date = None
        """날짜별로 각 방 타입의 판매가능객실 설정"""
        try:
            print("\n🏨 판매가능객실 자동 설정 중...")
            time.sleep(3)
            
            results = {}
            total_changed = 0  # 실제 변경된 입력 필드 수
            closed_rooms_list = []  # 마감된 방 추적
            
            for room_key, room_name in config.ROOM_TYPES.items():
                max_count = config.ROOM_MAX_COUNT.get(room_key, 10)
                print(f"\n📝 {room_name} 처리 중 (최대: {max_count}개)")
                
                try:
                    # 1. 먼저 룸 타입의 expandable 버튼 찾기 (여러 방법 시도)
                    expandable_span = None
                    try:
                        # 방법 1: class와 텍스트로 찾기
                        expandable_span = self.driver.find_element(
                            By.XPATH,
                            f"//span[contains(@class, 'expandable') and contains(text(), '{room_name}')]"
                        )
                    except:
                        # 방법 2: 아이콘 포함된 span으로 찾기
                        expandable_span = self.driver.find_element(
                            By.XPATH,
                            f"//span[@class='expandable' and contains(., '{room_name}')]"
                        )
                    
                    # 2. expandable 버튼 클릭하여 하위 메뉴 펼치기
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", expandable_span)
                    time.sleep(1)
                    
                    # 이미 펼쳐져 있는지 확인 (closes 아이콘이 보이는지)
                    parent_element = expandable_span.find_element(By.XPATH, "./ancestor::td")
                    closes_icon = parent_element.find_elements(By.CSS_SELECTOR, "i.closes")
                    
                    if not closes_icon or not closes_icon[0].is_displayed():
                        # 펼쳐져 있지 않으면 클릭
                        self.driver.execute_script("arguments[0].click();", expandable_span)
                        print(f"  ✓ {room_name} 하위 메뉴 펼침")
                        time.sleep(2)  # 하위 메뉴가 펼쳐질 때까지 대기
                    else:
                        print(f"  ✓ {room_name} 이미 펼쳐져 있음")
                    
                    # 3. 펼쳐진 하위 행들 중 "판매가능객실" 행의 입력 필드만 찾기
                    input_fields = []
                    try:
                        # expandable span이 있는 tr 찾기
                        expandable_tr = expandable_span.find_element(By.XPATH, "./ancestor::tr")
                        
                        # 다음 형제 tr들을 순회하면서 "판매가능객실"이 있는 tr 찾기
                        next_tr = expandable_tr
                        found_sales_row = False
                        
                        for i in range(20):  # 최대 20개 행 확인
                            try:
                                next_tr = next_tr.find_element(By.XPATH, "./following-sibling::tr[1]")
                                
                                # 다음 expandable을 만나면 중단 (다른 룸 타입 시작)
                                if next_tr.find_elements(By.CSS_SELECTOR, "span.expandable"):
                                    print(f"  → 다음 룸 타입 도달, 검색 중단")
                                    break
                                
                                # 이 tr에 "판매가능객실" 텍스트가 있는지 확인
                                tr_text = next_tr.text
                                if "판매가능객실" in tr_text:
                                    # 이 tr의 input 필드들 찾기
                                    row_inputs = next_tr.find_elements(By.CSS_SELECTOR, "input[type='text']")
                                    if row_inputs:
                                        input_fields = row_inputs
                                        found_sales_row = True
                                        print(f"  → {room_name}의 판매가능객실 행 발견: {len(input_fields)}개 입력 필드")
                                        break
                                
                            except:
                                break
                        
                        if not found_sales_row:
                            print(f"  ⚠ {room_name}의 판매가능객실 행을 찾을 수 없음")
                            results[room_name] = False
                            continue
                        
                    except Exception as e:
                        print(f"  ⚠ 판매가능객실 행 찾기 실패: {e}")
                        import traceback
                        traceback.print_exc()
                        results[room_name] = False
                        continue
                    
                    count = 0
                    alert_count = 0
                    skip_count = 0
                    # 잔여/예약 tr 찾기: 판매가능객실 tr의 이전 형제 tr 중 data-field="REMANING"인 tr
                    sales_tr = input_fields[0].find_element(By.XPATH, "./ancestor::tr")
                    remain_tr = sales_tr.find_element(By.XPATH, "./preceding-sibling::tr[@data-field='REMANING'][1]")
                    remain_tds = remain_tr.find_elements(By.TAG_NAME, "td")[2:]  # 앞 2개는 라벨/헤더
                    
                    # 마감 상태 확인 (판매가능객실 행 이전의 CLOSE_YN 행 찾기)
                    is_closed = False
                    try:
                        close_row = sales_tr.find_element(By.XPATH, "./preceding-sibling::tr[@data-field='CLOSE_YN'][1]")
                        close_text = close_row.text.strip().lower()
                        if '마감' in close_row.text or 'close' in close_text or '품절' in close_row.text:
                            is_closed = True
                            if room_name not in closed_rooms_list:
                                closed_rooms_list.append(room_name)
                            print(f"  ⊘ {room_name}: 마감 상태 감지 - 스킵")
                    except Exception:
                        pass

                    for idx, input_field in enumerate(input_fields):
                        try:
                            current_value = input_field.get_attribute('value')
                            remain_td = remain_tds[idx]
                            spans = remain_td.find_elements(By.TAG_NAME, "span")
                            remaining = int(spans[0].text.strip())
                            booked = int(spans[1].text.strip())
                            if idx < 3:
                                print(f"    [{idx+1}] 전여:{remaining}, 예약:{booked}")
                            print(f"[잔여/예약 진단] idx={idx+1}, room={room_key}, remaining={remaining}, booked={booked}, current_value={current_value}")
                        except Exception as e:
                            print(f"[잔여/예약 진단] idx={idx+1}, room={room_key}, remain_td 파싱 실패: {e}")
                            remaining = max_count
                            booked = 0
                        
                        try:
                            # 판매가능객실 수량 계산
                            available = self.calculate_available_rooms(
                                room_key, remaining, booked, max_count
                            )

                            # ALERT 메시지 처리
                            if isinstance(available, str) and available.startswith('ALERT:'):
                                alert_msg = available.replace('ALERT:', '')
                                print(f"    [{idx+1}] ⚠️ {alert_msg}")
                                alert_count += 1
                                continue

                            # ★ 빈칸→빈칸이면 완전 생략
                            if (current_value is None or str(current_value).strip() == "") and available is None:
                                skip_count += 1
                                if skip_count <= 3:
                                    print(f"    [{idx+1}] ✓ 건너뛰기: 빈칸→빈칸 (예약:{booked})")
                                count += 1
                                continue

                            # 값이 있고 조건에 맞으면 즉시 건너뛰기 (스크롤/클릭 없이)
                            # 기존 값과 기대값이 다르면 반드시 값을 입력하도록 수정
                            if current_value and str(current_value).strip():
                                try:
                                    existing_val = int(current_value)
                                    if available is not None and existing_val == available:
                                        skip_count += 1
                                        if skip_count <= 3:
                                            print(f"    [{idx+1}] ✓ 건너뛰기: {existing_val} (정책 기대값과 동일)")
                                        count += 1
                                        continue
                                except:
                                    pass

                            # 변경 이력 기록 (변경 발생 시)
                            if str(current_value) != ("" if available is None else str(available)):
                                # index에 따라 실제 날짜 계산
                                try:
                                    base_date = datetime.strptime(current_date, "%Y-%m-%d")
                                    real_date = (base_date + timedelta(days=idx)).strftime("%Y-%m-%d")
                                except Exception as e:
                                    real_date = current_date
                                self.change_history.append({
                                    'run_date': run_date,
                                    'date': real_date,
                                    'room_type': room_name,
                                    'old_value': current_value,
                                    'new_value': available
                                })
                            
                            # 입력 필드가 화면에 보이도록 스크롤
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_field)
                            time.sleep(0.5)
                            
                            # 입력 필드 활성화 및 포커스
                            self.driver.execute_script("arguments[0].removeAttribute('readonly');", input_field)
                            self.driver.execute_script("arguments[0].removeAttribute('disabled');", input_field)
                            input_field.click()
                            time.sleep(0.3)
                            
                            # 입력 필드에 값 설정 (여러 방법 시도)
                            if available is None:
                                # 빈칸 (모두 오픈)
                                try:
                                    input_field.clear()
                                except:
                                    self.driver.execute_script("arguments[0].value = '';", input_field)
                                print(f"    [{idx+1}] 빈칸으로 설정 (모두 오픈)")
                                total_changed += 1

                            else:
                                value_str = str(available)
                                success = False
                                for attempt in range(3):
                                    # 방법 1: clear + send_keys
                                    try:
                                        input_field.clear()
                                        time.sleep(0.2)
                                        input_field.send_keys(value_str)
                                        time.sleep(0.2)
                                    except Exception as e1:
                                        print(f"    방법1 실패, 방법2 시도: {e1}")
                                        # 방법 2: JavaScript로 직접 설정
                                        try:
                                            self.driver.execute_script(f"arguments[0].value = '';", input_field)
                                            time.sleep(0.1)
                                            self.driver.execute_script(f"arguments[0].value = '{value_str}';", input_field)
                                        except Exception as e2:
                                            print(f"    방법2도 실패: {e2}")
                                    # 변경 이벤트 트리거
                                    self.driver.execute_script("""
                                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                                    """, input_field)
                                    # blur 이벤트로 완료
                                    input_field.send_keys(Keys.TAB)
                                    time.sleep(0.2)
                                    # 입력값 검증
                                    actual_val = input_field.get_attribute('value')
                                    if actual_val == value_str:
                                        print(f"    [{idx+1}] 값 입력 성공: {value_str} (예약:{booked}, 잔여:{remaining})")
                                        success = True
                                        total_changed += 1
                                        break
                                    else:
                                        print(f"    [{idx+1}] 값 입력 불일치: 기대={value_str}, 실제={actual_val} (재시도 {attempt+1}/3)")
                                if not success:
                                    print(f"    [{idx+1}] ⚠️ 최종 입력 실패: {value_str} (예약:{booked}, 잔여:{remaining})")
                                count += 1
                            
                        except Exception as e:
                            print(f"  ⚠ 입력 필드 {idx+1} 설정 실패: {e}")
                            import traceback
                            traceback.print_exc()
                            continue
                    
                    print(f"  ✓ {room_name}: {count}개 처리 완료 (건너뛰기: {skip_count}개)")
                    if alert_count > 0:
                        print(f"  ⚠️ {room_name}: {alert_count}개 알림 - 수동 확인 필요")
                    results[room_name] = True
                    
                except Exception as e:
                    print(f"  ❌ {room_name} 설정 실패: {e}")
                    import traceback
                    traceback.print_exc()
                    results[room_name] = False
            
            # 저장 전 변경사항 체크
            if total_changed == 0:
                print("\n⏭️  변경사항이 없어서 저장을 건너뜁니다.")
                return results
            
            # 저장 버튼 클릭
            print(f"\n💾 저장 중... (변경된 입력 필드: {total_changed}개)")
            save_button = self.driver.find_element(
                By.CSS_SELECTOR,
                "#scrollArea > div:nth-child(1) > div.app-main > div.app-main__outer > div > div > div.app-footer.fixFooter.TabsAnimation-appear.TabsAnimation-appear-active > div > div > button.btn-wide.btn-shadow.w140.btn.btn-primary.btn-lg"
            )
            save_button.click()
            time.sleep(2)
            
            # "저장되었습니다" 팝업의 확인 버튼 클릭
            try:
                confirm_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-primary') and contains(., '확인')]"))
                )
                confirm_button.click()
                print("  ✓ 저장 확인 완료")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠ 확인 버튼 클릭 건너뜀: {e}")
            
            print("✅ 저장 완료")
            
            # 마감된 방이 있으면 하이라이트 로그 기록
            if closed_rooms_list:
                print(f"\n📊 마감된 방 {len(closed_rooms_list)}개에 대해 하이라이트 로그 기록 중...")
                self.highlight_closed_rooms_in_excel(current_date, closed_rooms_list)
            
            return results
            
        except Exception as e:
            print(f"❌ 판매가능객실 설정 실패: {e}")
            import traceback
            traceback.print_exc()
            return {}
        
    def set_room_availability(self, room_type, available_rooms):
        """
        특정 방 타입의 예약 가능 수량 설정
        
        Args:
            room_type (str): 방 타입 ('SINGLE', 'TWIN', 'DOUBLE', 'TRIPLE')
            available_rooms (int): 예약 가능한 방 수량
        """
        try:
            room_name = config.ROOM_TYPES.get(room_type)
            if not room_name:
                print(f"❌ 올바르지 않은 방 타입: {room_type}")
                return False
            
            print(f"\n🔄 {room_name} 예약 가능 수량 설정 중: {available_rooms}개")
            
            # 실제 CMS의 HTML 구조에 맞게 수정 필요
            # 예시: 방 타입별 입력 필드 찾기
            # XPath는 실제 페이지 구조를 확인 후 수정해야 합니다
            
            # 방법 1: 텍스트로 요소 찾기
            room_element = self.wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, f"//td[contains(text(), '{room_name}')]")
                )
            )
            
            # 해당 행에서 입력 필드 찾기
            # 실제 구조에 맞게 수정 필요
            input_field = room_element.find_element(
                By.XPATH, 
                ".//following::input[@type='number' or @type='text'][1]"
            )
            
            # 기존 값 지우고 새 값 입력
            input_field.clear()
            input_field.send_keys(str(available_rooms))
            
            print(f"✓ {room_name} 수량 설정 완료: {available_rooms}개")
            time.sleep(0.5)
            return True
            
        except Exception as e:
            print(f"❌ {room_name} 설정 실패: {e}")
            return False
    
    def set_all_rooms(self, room_settings):
        """
        모든 방 타입의 예약 가능 수량을 한번에 설정
        
        Args:
            room_settings (dict): {'SINGLE': 5, 'TWIN': 3, 'DOUBLE': 4, 'TRIPLE': 2}
        """
        print("\n" + "="*50)
        print("전체 방 예약 가능 수량 설정 시작")
        print("="*50)
        
        results = {}
        for room_type, count in room_settings.items():
            results[room_type] = self.set_room_availability(room_type, count)
        
        # 저장 버튼 클릭 (실제 구조에 맞게 수정)
        try:
            save_button = self.driver.find_element(
                By.XPATH, 
                "//button[contains(text(), '저장') or contains(text(), 'Save')]"
            )
            save_button.click()
            print("\n✓ 변경사항 저장 완료")
            time.sleep(2)
        except Exception as e:
            print(f"\n⚠ 저장 버튼을 찾을 수 없습니다: {e}")
            print("수동으로 저장해주세요.")
        
        # 결과 출력
        print("\n" + "="*50)
        print("설정 결과:")
        for room_type, success in results.items():
            status = "✓ 성공" if success else "❌ 실패"
            print(f"  {config.ROOM_TYPES[room_type]}: {status}")
        print("="*50)
        
        return results
    
    def get_current_availability(self):
        """현재 각 방의 예약 가능 수량 조회"""
        print("\n현재 예약 가능 수량 조회 중...")
        
        current_status = {}
        for room_type, room_name in config.ROOM_TYPES.items():
            try:
                # 실제 구조에 맞게 수정 필요
                room_element = self.driver.find_element(
                    By.XPATH, 
                    f"//td[contains(text(), '{room_name}')]"
                )
                input_field = room_element.find_element(
                    By.XPATH,
                    ".//following::input[@type='number' or @type='text'][1]"
                )
                current_value = input_field.get_attribute('value')
                current_status[room_type] = current_value
                print(f"  {room_name}: {current_value}개")
                
            except Exception as e:
                print(f"  {room_name}: 조회 실패 ({e})")
                current_status[room_type] = None
        
        return current_status
    
    def clear_all_room_availability(self):
        """모든 방 타입의 판매가능객실 입력값을 지우고 저장"""
        try:
            print("\n🧹 모든 판매가능객실 입력값 초기화 중...")
            time.sleep(3)
            
            results = {}
            
            for room_key, room_name in config.ROOM_TYPES.items():
                print(f"\n📝 {room_name} 초기화 중...")
                
                try:
                    # 1. expandable 버튼 찾기
                    expandable_span = None
                    try:
                        expandable_span = self.driver.find_element(
                            By.XPATH,
                            f"//span[contains(@class, 'expandable') and contains(text(), '{room_name}')]"
                        )
                    except:
                        expandable_span = self.driver.find_element(
                            By.XPATH,
                            f"//span[@class='expandable' and contains(., '{room_name}')]"
                        )
                    
                    # 2. 하위 메뉴 펼치기
                    self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", expandable_span)
                    time.sleep(1)
                    
                    parent_element = expandable_span.find_element(By.XPATH, "./ancestor::td")
                    closes_icon = parent_element.find_elements(By.CSS_SELECTOR, "i.closes")
                    
                    if not closes_icon or not closes_icon[0].is_displayed():
                        self.driver.execute_script("arguments[0].click();", expandable_span)
                        print(f"  ✓ {room_name} 하위 메뉴 펼침")
                        time.sleep(2)
                    else:
                        print(f"  ✓ {room_name} 이미 펼쳐져 있음")
                    
                    # 3. 판매가능객실 행의 입력 필드 찾기
                    input_fields = []
                    try:
                        sales_available_row = self.driver.find_element(
                            By.XPATH,
                            f"//span[contains(@class, 'expandable') and contains(text(), '{room_name}')]/ancestor::tr/following-sibling::tr[.//td[contains(text(), '판매가능객실')]]"
                        )
                        
                        input_fields = sales_available_row.find_elements(By.CSS_SELECTOR, "input[type='text']")
                        print(f"  → {len(input_fields)}개 입력 필드 찾음")
                        
                    except Exception as e:
                        print(f"  ⚠ 판매가능객실 행을 찾을 수 없음: {e}")
                        results[room_name] = False
                        continue
                    
                    # 4. 모든 입력값 지우기
                    count = 0
                    for input_field in input_fields:
                        try:
                            self.driver.execute_script("arguments[0].scrollIntoView({block: 'nearest'});", input_field)
                            time.sleep(0.1)
                            
                            # 값을 빈 문자열로 설정
                            self.driver.execute_script("arguments[0].value = '';", input_field)
                            self.driver.execute_script("arguments[0].focus(); arguments[0].blur();", input_field)
                            
                            count += 1
                            
                        except Exception as e:
                            continue
                    
                    print(f"  ✓ {room_name}: {count}개 입력값 초기화 완료")
                    results[room_name] = True
                    
                except Exception as e:
                    print(f"  ❌ {room_name} 초기화 실패: {e}")
                    results[room_name] = False
            
            # 저장 버튼 클릭
            print("\n💾 저장 중...")
            save_button = self.driver.find_element(
                By.CSS_SELECTOR,
                "#scrollArea > div:nth-child(1) > div.app-main > div.app-main__outer > div > div > div.app-footer.fixFooter.TabsAnimation-appear.TabsAnimation-appear-active > div > div > button.btn-wide.btn-shadow.w140.btn.btn-primary.btn-lg"
            )
            save_button.click()
            time.sleep(2)
            
            # "저장되었습니다" 팝업의 확인 버튼 클릭
            try:
                confirm_button = self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(@class, 'btn-primary') and contains(., '확인')]"))
                )
                confirm_button.click()
                print("  ✓ 저장 확인 완료")
                time.sleep(1)
            except Exception as e:
                print(f"  ⚠ 확인 버튼 클릭 건너뜀: {e}")
            
            print("✅ 초기화 및 저장 완료!")
            
            return results
            
        except Exception as e:
            print(f"❌ 초기화 실패: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def close(self):
        """브라우저 종료"""
        if self.driver:
            self.driver.quit()
            print("\n✓ 브라우저 종료")
    
    def run_for_date_range(self):
        """날짜 범위에 대해 자동 실행"""
        start_date = datetime(2026, 11, 1)
        end_date = datetime(2026, 12, 31)
        delta = timedelta(days=15)
        current_date = start_date
        while current_date <= end_date:
            date_str = current_date.strftime("%Y-%m-%d")
            print(f"\n===== {date_str} ~ 15일치 처리 시작 =====")
            self.set_date(date_str)
            self.set_room_availability_by_date()
            # 15일 뒤로 이동
            current_date += delta
    
    def run_for_date_range_with_input(self, start_date_str, end_date_str):
        """
        시작일(YYYY-MM-DD) 문자열을 받아 15일(시작일~시작일+14일)만 처리
        최초 1회만 객실/필터 설정, 이후에는 날짜만 바꾸고 반드시 조회 버튼을 누름
        """
        # 시작일 미입력 시 오늘+3일로 자동 설정
        if not start_date_str or start_date_str.strip() == "":
            start_date = datetime.now() + timedelta(days=3)
            start_date_str = start_date.strftime("%Y-%m-%d")
            print(f"시작일 미입력: 오늘+3일({start_date_str})로 자동 설정합니다.")
        else:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d")

        # 종료일 미입력 시 오늘+11개월로 자동 설정
        if not end_date_str or end_date_str.strip() == "":
            end_date = datetime.now() + timedelta(days=30*11)
            print(f"종료일 미입력: 오늘+11개월({end_date.strftime('%Y-%m-%d')})로 자동 설정합니다.")
        else:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
        current_date = start_date
        # 최초 1회: 객실/필터 설정 포함
        if current_date <= end_date:
            date_str = current_date.strftime("%Y-%m-%d")
            end_range = current_date + timedelta(days=14)
            print(f"\n===== {date_str} ~ {end_range.strftime('%Y-%m-%d')} 처리 시작 =====")
            self.navigate_to_inventory_page(date_str, do_select_rooms=True)
            self.set_room_availability_by_date()
            current_date += timedelta(days=15)
        # 이후 반복: 날짜만 바꾸고 조회 버튼 누름
        while current_date <= end_date:
            date_str = current_date.strftime("%Y-%m-%d")
            end_range = current_date + timedelta(days=14)
            print(f"\n===== {date_str} ~ {end_range.strftime('%Y-%m-%d')} 처리 시작 =====")
            self.set_date(date_str)
            self.search_rooms_by_date()
            self.set_room_availability_by_date()
            current_date += timedelta(days=15)


def test_highlight_closed_rooms():
    """마감 방 하이라이트 기능 테스트"""
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    print("\n" + "="*60)
    print("🧪 마감 방 하이라이트 테스트")
    print("="*60)
    
    excel_path = "기준가격.xlsx"
    
    try:
        # 테스트 데이터
        test_date = "2026-01-26"
        test_closed_rooms = ["Family Room 3 person", "Twin Room"]  # 테스트용 마감 방
        
        print(f"\n📋 테스트 설정:")
        print(f"  - 날짜: {test_date}")
        print(f"  - 마감 방: {test_closed_rooms}")
        
        # 엑셀 파일 열기
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 날짜 행 찾기
        target_row = None
        for row_idx, cell in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1), 1):
            try:
                cell_date = pd.Timestamp(cell[0].value).strftime("%Y-%m-%d")
                if cell_date == test_date:
                    target_row = row_idx
                    break
            except Exception:
                continue
        
        if target_row is None:
            print(f"❌ {test_date} 행을 찾지 못했습니다.")
            return False
        
        print(f"\n✓ {test_date} 행 발견 (행 번호: {target_row})")
        
        # 헤더 행의 컬럼명 확인 (디버깅용)
        print(f"\n📌 Excel 컬럼 정보:")
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            data = ws.cell(row=target_row, column=col_idx).value
            print(f"  Col {col_idx}: '{header}' = {data}")
        
        # 진한 노란색 적용
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        highlighted_count = 0
        
        print(f"\n🎨 하이라이트 적용 중:")
        for col_idx in range(1, ws.max_column + 1):
            room_type = ws.cell(row=1, column=col_idx).value
            if room_type in test_closed_rooms:
                cell = ws.cell(row=target_row, column=col_idx)
                cell.fill = yellow_fill
                highlighted_count += 1
                print(f"  ✓ '{room_type}' → 진한 노란색 적용")
        
        # 파일 저장
        wb.save(excel_path)
        
        print(f"\n✅ 테스트 완료!")
        print(f"  - 총 {highlighted_count}개 셀 하이라이트")
        print(f"  - 파일 저장: {excel_path}")
        return True
        
    except Exception as e:
        print(f"❌ 테스트 실패: {e}")
        import traceback
        traceback.print_exc()
        return False


def main():
    """메인 실행 함수"""
    import sys
    import io
    # Windows에서 UTF-8 출력 지원
    if sys.stdout.encoding != 'utf-8':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

    print("\n실행할 기능을 선택하세요:")
    print("1. 객실수 자동조정 (기간별)")
    print("2. 요금 자동입력 (RMO 기반)")
    print("3. 현재 요금 스냅샷 저장 (기준가격 형식)")
    option = input("번호 입력 (1~3): ").strip()

    # 기능별 입력값 미리 받기
    if option == "1":
        print("\n[객실수 자동조정] 기간을 입력하세요.")
        start_date_str = input("시작일 (YYYY-MM-DD): ")
        end_date_str = input("종료일 (YYYY-MM-DD): ")
    elif option == "2":
        print("\n[요금 자동입력] 기간을 입력하세요.")
        start_date_str = input("시작일 (YYYY-MM-DD, 엔터시 오늘): ")
        end_date_str = input("종료일 (YYYY-MM-DD, 엔터시 시작일+14일): ")
    elif option == "3":
        print("\n[현재 요금 스냅샷] 기간을 입력하세요.")
        start_date_str = input("시작일 (YYYY-MM-DD, 엔터시 오늘): ")
        end_date_str = input("종료일 (YYYY-MM-DD, 엔터시 시작일+14일): ")
    else:
        start_date_str = None
        end_date_str = None

    controller = HotelCMSController()

    try:
        controller.setup_driver()
        controller.navigate_to_cms()
        login_success = controller.login()
        if not login_success:
            print("\n수동으로 로그인을 완료한 후 Enter를 눌러주세요...")
            input()

        if option == "1":
            controller.run_for_date_range_with_input(start_date_str, end_date_str)
            # 변경 이력 엑셀로 저장
            if controller.change_history:
                import pandas as pd
                df = pd.DataFrame(controller.change_history)
                df.to_excel("change_history.xlsx", index=False)
                print(f"\n변경 이력(change_history.xlsx) 저장 완료! 변경 건수: {len(df)}")
            else:
                print("\n변경된 내역이 없습니다.")
            print("\n" + "="*60)
            print("✅ 기간별 판매가능객실 설정 완료!")
            print("="*60)
        elif option == "2":
            # 미입력시 오늘 날짜로 자동
            if not start_date_str or start_date_str.strip() == "":
                start_date_str = datetime.now().strftime("%Y-%m-%d")
                print(f"시작일 미입력: 오늘({start_date_str})로 자동 설정합니다.")
            
            # 종료일 미입력시 시작일+14일로 자동
            if not end_date_str or end_date_str.strip() == "":
                start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
                end_date = start_dt + timedelta(days=14)
                end_date_str = end_date.strftime("%Y-%m-%d")
                print(f"종료일 미입력: 시작일+14일({end_date_str})로 자동 설정합니다.")
            else:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            
            # 시작일부터 종료일까지 15일씩 반복 처리 (매번 +14일 화면에서 한 번에 처리)
            current_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            
            while current_date <= end_date:
                date_str = current_date.strftime("%Y-%m-%d")
                next_date = current_date + timedelta(days=15)
                print(f"\n📅 {date_str}부터 +14일 처리 중...")
                controller.auto_set_rates_by_rmo(date_str)
                current_date = next_date
            
            print("\n" + "="*60)
            print("✅ 요금 자동입력 완료!")
            print("="*60)
        elif option == "3":
            # 미입력시 1월 26일로 자동 설정
            if not start_date_str or start_date_str.strip() == "":
                start_date_str = "2026-01-26"
                print(f"시작일 미입력: 1월 26일({start_date_str})로 자동 설정합니다.")
            # 종료일 미입력시 시작일+14일로 자동
            if not end_date_str or end_date_str.strip() == "":
                start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
                end_date = start_dt + timedelta(days=14)
                end_date_str = end_date.strftime("%Y-%m-%d")
                print(f"종료일 미입력: 시작일+14일({end_date_str})로 자동 설정합니다.")
            else:
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d")

            current_date = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            print("\n📸 현재 요금 스냅샷 수집 시작...")
            while current_date <= end_date:
                date_str = current_date.strftime("%Y-%m-%d")
                print(f"\n📅 {date_str}부터 +14일 스냅샷 수집...")
                controller.snapshot_current_rates_by_rmo(date_str)
                current_date = current_date + timedelta(days=15)
            controller.save_snapshot_to_excel()
            print("\n" + "="*60)
            print("✅ 현재 요금 스냅샷 저장 완료!")
            print("="*60)
        else:
            print("잘못된 옵션입니다. 프로그램을 종료합니다.")

    except KeyboardInterrupt:
        print("\n\n사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
    finally:
        if controller and getattr(controller, "driver", None):
            try:
                input("\n모든 처리가 끝났습니다. 창을 닫으려면 엔터를 누르세요... ")
            except Exception:
                pass
        controller.close()


if __name__ == "__main__":
    main()
